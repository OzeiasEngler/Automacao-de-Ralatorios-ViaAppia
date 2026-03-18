r"""
Gera «Exportar Kcor.xlsx» (lote 50 Artemig).
V: base\_02 Arquivos Fotos + subpasta por apontamento (ex.: NOT-25-01365_PAVIMENTO_CE2516929).
W: PDF (COD).jpg; nc (COD).jpg; nc (COD)_N.jpg — ficheiros dentro dessa subpasta.
"""
from __future__ import annotations

import io
import logging
import os
import re
import shutil
import tempfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def _aplicar_bordas_linha_kcor(ws, row: int, col_fim: int = 25) -> None:
    from openpyxl.styles import Border, Side

    b = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for col in range(1, col_fim + 1):
        ws.cell(row=row, column=col).border = b


def _desfazer_merge_colunas_linha_kcor(ws, row: int, col_ini: int, col_fim: int) -> None:
    from openpyxl.utils.cell import range_boundaries

    to_unmerge = []
    for mc in list(ws.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(mc))
        except Exception:
            continue
        if row >= min_row and row <= max_row and not (col_fim < min_col or col_ini > max_col):
            to_unmerge.append(mc)
    for mc in to_unmerge:
        try:
            ws.unmerge_cells(str(mc))
        except Exception:
            pass


def _copiar_estilo_linha_kcor(ws, row_origem: int, row_destino: int, col_fim: int = 25) -> None:
    col_ate_u = min(21, col_fim)
    for col in range(1, col_ate_u + 1):
        src = ws.cell(row=row_origem, column=col)
        dst = ws.cell(row=row_destino, column=col)
        if src.has_style:
            dst.font = src.font.copy()
            dst.border = src.border.copy()
            dst.fill = src.fill.copy()
            dst.number_format = src.number_format
            dst.alignment = src.alignment.copy()
    _aplicar_bordas_linha_kcor(ws, row_destino, col_fim)

_CLASS = "Eng. QID"


def _patologia_para_kcor(pat: str, indicador: str, atividade: str) -> tuple[str, str]:
    s = f"{pat} {indicador} {atividade}".lower()
    rules: list[tuple[str, str]] = [
        ("buraco", "Buracos e panelas - Emergencial "),
        ("panela", "Buracos e panelas - Emergencial "),
        ("trilha", "Afundamento nas trilhas de rodas"),
        ("alambrado", "Alambrado"),
        ("dispositivo de segurança", "Alambrado"),
        ("guarda corpo", "Barreira rígida "),
        ("inexistência de elementos refletivos", "Barreira rígida "),
        ("caiação", "Caiação"),
        ("caiacao", "Caiação"),
        ("cerca", "Cerca"),
        ("erosão", "Conservação de terraplenos e contenções"),
        ("erosao", "Conservação de terraplenos e contenções"),
        ("defensa", "Defensa metálica"),
        ("deformação", "Deformação permanente "),
        ("deformacao", "Deformação permanente "),
        ("degrau", "Degrau em acostamento"),
        ("sinalização vertical", "Demais placas"),
        ("sinalizacao vertical", "Demais placas"),
        ("demais placas", "Demais placas"),
        ("vandalismo", "Demais placas"),
        ("iluminação", "Dispositivos de Iluminação"),
        ("iluminacao", "Dispositivos de Iluminação"),
        ("drenagem subterrânea", "Drenagem Subterrânea"),
        ("drenagem subterranea", "Drenagem Subterrânea"),
        ("drenagem", "Drenagem Superficial"),
        ("entulho", "Entulho"),
        ("horizontal", "Sinalização horizontal"),
        ("tacha", "Tachas e tachões"),
        ("tachao", "Tachas e tachões"),
        ("vegetação", "Vegetação"),
        ("vegetacao", "Vegetação"),
    ]
    for kw, kcor in rules:
        if kw in s:
            return kcor.rstrip() + (" " if kcor.endswith(" ") else ""), _CLASS
    if "buraco" in s or "panela" in s:
        return "Buracos e panelas - Reparo técnico", _CLASS
    p0 = (pat or "").strip()[:120]
    return (p0 or "Patologia — conferir mapeamento Kcor"), _CLASS


def _parse_dt(s: str) -> datetime | None:
    t = (s or "").strip()
    if " " in t:
        t = t.split()[0].strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(t, fmt)
        except ValueError:
            continue
    return None


def _origem(tipo: str) -> str:
    t = (tipo or "").strip().upper()
    if "QID" in t:
        return "0-QID"
    return "Orgão Fiscalizador"


def _prazo_dias_efetivo(nc: Any) -> int:
    d = getattr(nc, "prazo_dias", None)
    if d is None:
        return 0
    try:
        n = int(d)
    except (TypeError, ValueError):
        return 0
    if n == 24 and getattr(nc, "emergencial", False):
        return 1
    return max(0, n)


def _codigo_fiscalizacao_arquivos(nc: Any) -> str:
    """Código da fiscalização (ex.: 202506784) — nomes PDF (COD).jpg / nc (COD).jpg."""
    c = (getattr(nc, "codigo", None) or "").strip()
    if not c:
        return ""
    if re.fullmatch(r"\d{6,14}", c):
        return c
    m = re.search(r"\b(\d{8,10})\b", c)
    return m.group(1) if m else c


def _rodovia_coluna_f(rod: str) -> str:
    r = re.sub(r"\s+", " ", (rod or "").strip().upper()).replace("-", " ")
    m = re.match(r"^(MG|BR)\s+(\d+)$", r)
    if m:
        pref, num = m.group(1), int(m.group(2))
        return f"{pref}-{num:03d}"
    if " " in r:
        return r.replace(" ", "-", 1)
    return (rod or "").strip()


def _local_coluna_j(nc: Any) -> str:
    blob = f"{nc.atividade or ''} {nc.tipo_atividade or ''} {nc.grupo_atividade or ''}".upper()
    if "DOM" in blob and "NIO" in blob or "FAIXA DE DOM" in blob or "FX." in blob:
        return "Faixa de Domínio"
    return "Faixa de Rolamento"


def _data_kcor_so_data(nc: Any) -> tuple[str, datetime.date | None]:
    """Data Kcor como dd/mm/aaaa (hora fica na coluna Hora)."""
    dt = _parse_dt(nc.data_con or "")
    if not dt:
        return "", None
    d = dt.date()
    return d.strftime("%d/%m/%Y"), d


def _stem_subpasta_fotos(nc: Any) -> str:
    """Pasta por NC: NOT-yy-xxxxx_PAVIMENTO_CE{consol} se houver código+consol; senão stem do PDF."""
    cod = _codigo_fiscalizacao_arquivos(nc)
    cons = (getattr(nc, "num_consol", None) or "").strip()
    if len(cod) >= 9 and cons.isdigit():
        yy = cod[2:4]
        seq = cod[4:9]
        return f"NOT-{yy}-{seq}_PAVIMENTO_CE{cons}"
    return (getattr(nc, "artemig_pdf_stem", None) or "").strip()


def _montar_v_w_kcor(nc: Any) -> tuple[str, str]:
    """Caminho pasta (V) e lista de JPG (W) por código de fiscalização da linha."""
    from nc_artemig.config import DIR_BASE_FOTOS_KCOR

    base = (DIR_BASE_FOTOS_KCOR or os.environ.get("ARTEMIG_KCOR_DIR_FOTOS") or "").strip()
    stem = _stem_subpasta_fotos(nc)
    cod = _codigo_fiscalizacao_arquivos(nc)
    pags = list(getattr(nc, "artemig_kcor_paginas_jpg", None) or [])

    if base and stem:
        v = os.path.normpath(os.path.join(base, stem))
    elif base:
        v = os.path.normpath(base)
    else:
        v = ""

    if not cod:
        return v, ""

    n_nc = max(1, len(pags)) if pags else 1
    w_parts = [f"PDF ({cod}).jpg", f"nc ({cod}).jpg"]
    for i in range(1, n_nc):
        w_parts.append(f"nc ({cod})_{i}.jpg")
    return v, ";".join(w_parts)


def _ordenar_ncs_por_codigo_kcor(ncs: list[Any]) -> list[Any]:
    """Uma linha por fiscalização; ordem estável por número do código (ligação Excel ↔ ficheiros)."""

    def chave(nc: Any) -> tuple:
        c = _codigo_fiscalizacao_arquivos(nc) or ""
        try:
            n = int(c) if c.isdigit() else 0
        except ValueError:
            n = 0
        return (n, nc.km_ini or 0.0, (nc.rodovia or ""), c)

    return sorted(ncs, key=chave)


def gerar_exportar_kcor_xlsx_bytes(ncs: list[Any]) -> bytes | None:
    from nc_artemig.config import COL_KCOR_KRIA, MODELO_KCOR_KRIA

    ncs50 = [n for n in ncs if (getattr(n, "lote", None) or "").strip() == "50"]
    if not ncs50:
        return None
    ncs50 = _ordenar_ncs_por_codigo_kcor(ncs50)
    modelo = Path(MODELO_KCOR_KRIA)
    if not modelo.is_file():
        logger.error("exportar_kcor: modelo inexistente %s", modelo)
        return None
    try:
        import openpyxl
    except ImportError:
        logger.error("exportar_kcor: openpyxl necessário")
        return None

    c = COL_KCOR_KRIA
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    try:
        os.close(fd)
        shutil.copy2(str(modelo), tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.active

        from openpyxl.styles import Border, Side

        _side_k = Side(style="thin", color="000000")
        _border_linha_k = Border(
            left=_side_k, right=_side_k, top=_side_k, bottom=_side_k
        )

        lin_mod = 2
        n_lin = len(ncs50)
        for r in range(lin_mod + 1, lin_mod + n_lin):
            if r > ws.max_row:
                ws.insert_rows(r, 1)
                _copiar_estilo_linha_kcor(ws, lin_mod, r, 25)
        for r in range(lin_mod, lin_mod + n_lin):
            _desfazer_merge_colunas_linha_kcor(ws, r, 17, 20)
            for col in range(1, 26):
                ws.cell(r, col).value = None

        for idx, nc in enumerate(ncs50, start=1):
            r = idx + 1
            cod_linha = _codigo_fiscalizacao_arquivos(nc)
            if not cod_linha:
                logger.warning(
                    "Exportar Kcor linha %s: sem código fiscalização; col. W vazia para esta NC",
                    idx,
                )
            pat = (getattr(nc, "patologia_artemig", None) or "") or (nc.grupo_atividade or "")
            ind = (getattr(nc, "indicador_artemig", None) or "") or ""
            kcor, classe = _patologia_para_kcor(pat, ind, nc.atividade or "")

            ws.cell(r, c["NumItem"], idx)
            ws.cell(r, c["Origem"], _origem(nc.tipo_artemig))
            ws.cell(r, c["Motivo"], "Conservação de Rotina")
            ws.cell(r, c["Classificacao"], classe)
            ws.cell(r, c["Tipo"], kcor)
            ws.cell(r, c["Rodovia"], _rodovia_coluna_f(nc.rodovia or ""))
            g = nc.km_ini if nc.km_ini is not None else _km_f(nc.km_ini_str)
            h = nc.km_fim if nc.km_fim is not None else g
            ws.cell(r, c["KMi"], g if g is not None else "")
            ws.cell(r, c["KMf"], h if h is not None else "")
            ws.cell(r, c["Sentido"], (nc.sentido or "").strip())
            ws.cell(r, c["Local"], _local_coluna_j(nc))
            ws.cell(r, c["Gestor"], "")
            ws.cell(r, c["Executores"], "")

            ds, d0 = _data_kcor_so_data(nc)
            pd = _prazo_dias_efetivo(nc)
            if ds and d0:
                ws.cell(r, c["Data_Solicitacao"], ds)
                ws.cell(r, c["Dt_Inicio_Prog"], ds)
                ws.cell(r, c["Dt_Inicio_Exec"], ds)
                if pd and getattr(nc, "emergencial", False):
                    fim = d0
                elif pd:
                    fim = d0 + timedelta(days=pd)
                else:
                    fim = d0
                ws.cell(r, c["Dt_Fim_Prog"], fim.strftime("%d/%m/%Y"))
                if pd:
                    ws.cell(r, c["Data_Suspensao"], (d0 + timedelta(days=pd)).strftime("%d/%m/%Y"))
            if pd:
                ws.cell(r, c["Prazo"], pd)
            elif getattr(nc, "prazo_dias", None) is not None:
                try:
                    ws.cell(r, c["Prazo"], int(nc.prazo_dias))
                except (TypeError, ValueError):
                    pass

            sh = (getattr(nc, "sh_artemig", None) or "").strip()
            og: list[str] = []
            if sh:
                og.append(f"Trecho Homogênio: {sh}")
            og.append(f"Notificação: {(nc.codigo or '').strip()}")
            if (getattr(nc, "num_consol", None) or "").strip():
                og.append(f"Nº Consol: {nc.num_consol.strip()}")
            ws.cell(r, c["Obs_Gestor"], "\n".join(og))

            u1 = f"{pat} ({ind})".strip(" ()") if pat or ind else ""
            u2 = (nc.atividade or "").strip()[:450]
            ws.cell(r, c["Observacoes"], "\n\n".join(x for x in (u1, u2) if x).strip()[:2000])

            v_dir, w_arq = _montar_v_w_kcor(nc)
            ws.cell(r, c["Diretorio"], v_dir)
            ws.cell(r, c["Arquivos"], w_arq)
            ws.cell(r, c["Indicador"], ind[:120] if ind else "")
            ws.cell(r, c["Unidade"], "")

            for col_k in (
                c["Data_Solicitacao"],
                c["Data_Suspensao"],
                c["Dt_Inicio_Prog"],
                c["Dt_Fim_Prog"],
                c["Dt_Inicio_Exec"],
                c["Dt_Fim_Exec"],
            ):
                cl = ws.cell(r, col_k)
                if cl.value is not None and str(cl.value).strip():
                    sv = str(cl.value).strip()
                    if " " in sv and re.match(r"^\d{1,2}/\d{1,2}/\d{4}", sv):
                        cl.value = sv.split()[0][:10]
                cl.number_format = "@"
            for col_k in range(1, 26):
                ws.cell(r, col_k).border = _border_linha_k

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception as e:
        logger.exception("exportar_kcor: %s", e)
        return None
    finally:
        try:
            Path(tmp).unlink(missing_ok=True)
        except OSError:
            pass


def _km_f(s: str) -> float | None:
    if not s:
        return None
    m = re.match(r"(\d+)\s*\+\s*(\d+)", str(s).strip())
    if m:
        return int(m.group(1)) + int(m.group(2)) / 1000.0
    try:
        return float(str(s).replace(",", "."))
    except ValueError:
        return None
