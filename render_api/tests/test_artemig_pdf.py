"""nc_artemig: normalização de texto PDF e datas Artemig."""

from types import SimpleNamespace


def test_excel_complemento_nao_mescla_campos_texto_kcor_lote50():
    from nc_artesp.modulos.analisar_pdf_nc import _excel_complemento_pode_mesclar_campo

    assert _excel_complemento_pode_mesclar_campo("50", "tipo_atividade") is False
    assert _excel_complemento_pode_mesclar_campo("50", "atividade") is False
    assert _excel_complemento_pode_mesclar_campo("50", "prazo_str") is False
    assert _excel_complemento_pode_mesclar_campo("50", "km_ini_str") is False
    assert _excel_complemento_pode_mesclar_campo("50", "km_fim_str") is False
    assert _excel_complemento_pode_mesclar_campo("50", "horario_fiscalizacao") is True
    assert _excel_complemento_pode_mesclar_campo("13", "tipo_atividade") is True


def test_normalizar_texto_extraido_pdf_nbsp_zw():
    from nc_artemig.texto_pdf import normalizar_texto_extraido_pdf

    raw = "MG\u00a0050\u200bSH02"
    out = normalizar_texto_extraido_pdf(raw)
    assert "\u00a0" not in out
    assert "\u200b" not in out
    assert "MG" in out and "050" in out and "SH02" in out


def test_data_artemig_dd_mm_yyyy_com_ruido():
    from nc_artesp.modulos.analisar_pdf_nc import _data_artemig_dd_mm_yyyy

    assert _data_artemig_dd_mm_yyyy("15/03/26 ") == "15/03/2026"
    assert _data_artemig_dd_mm_yyyy("Data \u00a0 01/12/2025 fim") == "01/12/2025"


def test_local_coluna_j_faixa_dominio():
    from nc_artemig.exportar_kcor_planilha import _local_coluna_j

    nc = SimpleNamespace(atividade="Trecho FAIXA DE DOMINIO", tipo_atividade="", grupo_atividade="")
    assert _local_coluna_j(nc) == "Faixa de Domínio"
    nc2 = SimpleNamespace(atividade="FX. marginal", tipo_atividade="", grupo_atividade="")
    assert _local_coluna_j(nc2) == "Faixa de Domínio"
    nc3 = SimpleNamespace(atividade="Pista rolamento", tipo_atividade="", grupo_atividade="")
    assert _local_coluna_j(nc3) == "Faixa de Rolamento"


def test_patologia_para_kcor_macro_inexistencia_defensa():
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    k, cl = _patologia_para_kcor("Inexistência de elementos refletivos", "", "")
    assert "Defensa" in k
    assert cl == "Eng. QID"


def test_patologia_para_kcor_macro_guarda_corpo():
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    k, _ = _patologia_para_kcor("Guarda corpo metálico danificado", "", "")
    assert "Barreira" in k


def test_patologia_para_kcor_macro_placas_advertencia():
    from nc_artemig.exportar_kcor_planilha import _patologia_para_kcor

    k, _ = _patologia_para_kcor("Vandalismo placas de advertência na via", "", "")
    assert "Placas - Regulam" in k


def test_km_normalizado_nas01_nao_divide_km_decimal_artemig():
    """653,4 km não pode virar 0,6534 (regra antiga «>500 → /1000»)."""
    from nc_artemig.exportar_kcor_planilha import _km_normalizado_nas01

    assert _km_normalizado_nas01(653.4) == 653.4
    assert _km_normalizado_nas01(653.0) == 653.0
    assert _km_normalizado_nas01(653400) == 653.4
    assert _km_normalizado_nas01(12000) == 12.0


def test_export_kcor_col_p_dt_fim_prog_apos_contagem_emergencial():
    import io
    from types import SimpleNamespace

    from openpyxl import load_workbook

    from nc_artemig.config import COL_KCOR_KRIA
    from nc_artemig.exportar_kcor_planilha import gerar_exportar_kcor_xlsx_bytes

    nc = SimpleNamespace(
        lote="50",
        codigo="202603080000",
        data_con="10/04/2026",
        prazo_dias=24,
        emergencial=True,
        tipo_panela=False,
        km_ini=100.0,
        km_fim=100.0,
        km_ini_str="100,000",
        km_fim_str="100,000",
        rodovia="MG-050",
        sentido="Crescente",
        atividade="Buracos e/ou panelas",
        tipo_atividade="Panelas",
        grupo_atividade="Pavimento",
        observacao="",
        num_consol="2607782",
        sh_artemig="SH02",
        patologia_artemig="Buracos e/ou panelas na pista",
        indicador_artemig="Buracos / Panelas",
        artemig_pdf_stem="X",
        artemig_kcor_nomes_arquivos=[],
        artemig_kcor_paginas_jpg=[],
        nome_fiscal="",
        prazo_str="11/04/2026",
    )
    b, meta = gerar_exportar_kcor_xlsx_bytes([nc])
    assert meta.get("ok") and b
    wb = load_workbook(io.BytesIO(b))
    ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.active
    ck = COL_KCOR_KRIA
    r = 2
    ini = str(ws.cell(r, ck["Dt_Inicio_Prog"]).value or "").strip()
    fim = str(ws.cell(r, ck["Dt_Fim_Prog"]).value or "").strip()
    assert ini.startswith("10/04/2026")
    assert fim.startswith("11/04/2026")
    assert fim.split()[0] != ini.split()[0]
    prazo_cell = ws.cell(r, ck["Prazo"])
    assert prazo_cell.alignment is not None
    assert prazo_cell.alignment.horizontal == "center"


def test_stem_subpasta_fotos_col_v_usa_stem_pdf_nao_pavimento_fixo():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _stem_subpasta_fotos

    nc = SimpleNamespace(
        codigo="202506768",
        num_consol="2516934",
        artemig_pdf_stem="NOT-25-06768_DRENAGEM_CE2516934",
        grupo_atividade="",
        tipo_atividade="",
        indicador_artemig="",
        patologia_artemig="",
        atividade="",
    )
    s = _stem_subpasta_fotos(nc)
    assert "DRENAGEM" in s
    assert "PAVIMENTO" not in s


def test_montar_col_w_lista_nomes_multiplos_jpg():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _montar_v_w_kcor

    nc = SimpleNamespace(
        codigo="202506999",
        num_consol="",
        emergencial=False,
        prazo_dias=None,
        artemig_pdf_stem="Relato1",
        artemig_kcor_nomes_arquivos=[
            "nc (202506999).jpg",
            "nc (202506999)_1.jpg",
            "nc (202506999)_2.jpg",
        ],
        artemig_kcor_paginas_jpg=[],
        atividade="",
        tipo_atividade="",
        grupo_atividade="",
    )
    _v, w = _montar_v_w_kcor(nc)
    assert "nc (202506999).jpg" in w
    assert "nc (202506999)_1.jpg" in w
    assert "nc (202506999)_2.jpg" in w
    assert w.count(";") >= 3


def test_prazo_artemig_em_ate_24_horas_com_parenteses():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: Remendo Emergencial: "
        "em até 24 (vinte e quatro) horas, a partir da data do recebimento desta notificação."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "10/04/2026")
    assert prazo_dias == 1
    assert emerg is True
    assert prazo_str == "11/04/2026"


def test_prazo_artemig_em_ate_ndias_com_parenteses_extenso():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "em até 30 (trinta) dias corridos, a partir da data do recebimento desta notificação."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "01/04/2026")
    assert prazo_dias == 30
    assert emerg is False
    assert prazo_str == "01/05/2026"


def test_prazo_artemig_prazo_maximo_de_ndias_na_janela():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "prazo máximo de 60 (sessenta) dias corridos."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "01/01/2026")
    assert prazo_dias == 60
    assert emerg is False
    assert prazo_str == "02/03/2026"


def test_prazo_artemig_em_at_ndias_sem_letra_apos_at():
    """PDFs/encodings em que «até» perde o «e»/«é» («Em at 5 dias»)."""
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "Em at 5 (cinco) dias, a partir da data do recebimento desta notificação."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "01/04/2026")
    assert prazo_dias == 5
    assert emerg is False
    assert prazo_str == "06/04/2026"


def test_prazo_artemig_buracos_so_primeiro_prazo_emergencial_ignora_tecnico():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "Remendo Emergencial: em até 24 (vinte e quatro) horas, a partir da data.\n"
        "Remendo Técnico: no prazo máximo de 15 (quinze) dias, a partir da data."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "15/09/2025")
    assert prazo_dias == 1
    assert emerg is True
    assert prazo_str == "16/09/2025"


def test_prazo_artemig_solo_remendo_tecnico_15_dias():
    from nc_artesp.modulos.analisar_pdf_nc import _prazo_artemig

    texto = (
        "Prazo para Atendimento à Notificação: "
        "Remendo Técnico: no prazo máximo de 15 (quinze) dias, a partir da data."
    )
    prazo_str, prazo_dias, emerg = _prazo_artemig(texto, "15/09/2025")
    assert prazo_dias == 15
    assert emerg is False
    assert prazo_str == "30/09/2025"


def test_indicador_patologia_drenagem_subterranea_patologia_completa():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = "Drenagem Subterrânea Drenagem subterrânea obstruída"
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Subterr" in ind
    assert "obstru" in pat.lower()


def test_indicador_patologia_parametros_gerais_patologia_completa():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = "Parâmetros Gerais\nInexistência de tachas e tachões"
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Parâmetros" in ind and "Gerais" in ind
    assert "Inexistência" in pat
    assert "tachas" in pat.lower()


def test_indicador_patologia_gerais_parametros_ordem_colunas_pdf():
    """PDF linearizado: «Gerais» + «Parâmetros» (ordem de colunas) não vira só «Gerais (Parâmetros)»."""
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = "Gerais Parâmetros Inexistência de sinalização vertical"
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Parâmetros" in ind and "Gerais" in ind
    assert "Inexistência" in pat
    assert "sinalização vertical" in pat.lower()
    assert "Gerais (Parâmetros)" not in pat


def test_indicador_patologia_resto_multilinha_antes_de_e_ou():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    resto = (
        "Buracos / Panelas e \n"
        "Deformação permanente\n"
        "Buracos e/ou panelas na pista de \n"
        "rolamento"
    )
    ind, pat = _indicador_patologia_de_resto_artemig(resto)
    assert "Deformação permanente" in ind
    assert "Panelas e" in ind
    assert "e/ou" in pat
    assert "rolamento" in pat


def test_indicador_patologia_resto_buracos_e_ou_colado():
    from nc_artesp.modulos.analisar_pdf_nc import _indicador_patologia_de_resto_artemig

    r = (
        "Buracos / Panelas e Deformação permanente "
        "Buracos e/ou panelas na pista de rolamento."
    )
    ind, pat = _indicador_patologia_de_resto_artemig(r)
    assert "Deformação permanente" in ind
    assert "e/ou" in pat
    assert "pista de rolamento" in pat


def test_col_u_enriquece_patologia_curta_com_tipo_planilha():
    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="Panelas e buracos emergencial na pista",
        grupo_atividade="Buracos / Panelas e Deformação permanente",
        indicador_artemig="Buracos",
        atividade="",
        observacao="",
        codigo="202603080",
        num_consol="2607782",
        sh_artemig="SH02",
    )
    u = _texto_observacoes_nas01(nc)
    assert "Panelas e buracos emergencial na pista" in u
    assert "Buracos / Panelas e Deformação permanente" in u
    assert "Patologia:" not in u


def test_col_u_patologia_panelas_e_completa_so_com_tipo_pdf_sem_atividade():
    """NC real: patologia PDF curta + tipo (mesmo PDF) longo; atividade pode vir vazia."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="Panelas e buracos emergencial na pista",
        grupo_atividade="Buracos / Panelas e Deformação permanente",
        indicador_artemig="Buracos",
        atividade="",
        observacao="",
        codigo="202501364",
        num_consol="2516928",
        sh_artemig="SH18",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202501364" in u and "SH18" in u and "2516928" in u
    assert "Panelas e buracos emergencial na pista" in u


def test_col_u_patologia_completa_desde_atividade_quando_pdf_truncado():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="",
        grupo_atividade="",
        indicador_artemig="Buracos",
        atividade="Buracos e/ou panelas na pista de rolamento",
        observacao="",
        codigo="202501364",
        num_consol="2516928",
        sh_artemig="SH18",
    )
    u = _texto_observacoes_nas01(nc)
    assert "Buracos e/ou panelas na pista de rolamento" in u
    assert "Panelas e." not in u
    assert "Notificação:" not in u and "Indicador:" not in u and "Patologia:" not in u


def test_col_u_nao_duplica_indicador_patologia_iguais_gerais():
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    g = "Gerais (Parâmetros)"
    nc = SimpleNamespace(
        patologia_artemig=g,
        tipo_atividade="",
        grupo_atividade=g,
        indicador_artemig=g,
        atividade="",
        observacao="",
        codigo="202501363",
        num_consol="2516927",
        sh_artemig="SH05",
    )
    u = _texto_observacoes_nas01(nc)
    assert u.count("Gerais (Parâmetros)") == 1
    assert "Gerais (Parâmetros)" in u
    assert "Notificação:" not in u and "Indicador:" not in u and "Patologia:" not in u


def test_col_u_patologia_panelas_e_completa_com_tipo_curto_mais_3_chars():
    """Antes exigia len(tipo) > len(pat)+14; «Panelas e buracos» não substituía «Panelas e»."""
    from types import SimpleNamespace

    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="Panelas e",
        tipo_atividade="Panelas e buracos",
        grupo_atividade="Buracos",
        indicador_artemig="Buracos",
        atividade="",
        observacao="",
        codigo="202501365",
        num_consol="2516929",
        sh_artemig="SH20",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202501365" in u and "SH20" in u and "2516929" in u
    assert "Panelas e buracos" in u


def test_col_u_observacoes_usa_patologia_tipo_excel_nao_grupo_indicador():
    """Col. U: rótulos PDF — Indicador (grupo Excel) antes de Patologia (tipo), sem confundir os campos."""
    from nc_artemig.exportar_kcor_planilha import _texto_observacoes_nas01

    nc = SimpleNamespace(
        patologia_artemig="",
        tipo_atividade="Panelas e buracos emergencial",
        grupo_atividade="Pavimento",
        indicador_artemig="",
        atividade="Limpeza da área",
        observacao="",
        codigo="202506123456",
        num_consol="987654",
        sh_artemig="SH02",
    )
    u = _texto_observacoes_nas01(nc)
    assert "202506123456" in u
    assert "SH02" in u
    assert "Pavimento" in u
    assert "Panelas e buracos emergencial" in u
    assert "Limpeza da área" in u
    assert "987654" in u
    assert "Notificação:" not in u and "SH:" not in u and "Patologia:" not in u
    assert u.index("Pavimento") < u.index("Panelas")


def test_parse_artemig_texto_km_inicial_final_rotulos_pdf():
    from nc_artesp.modulos.analisar_pdf_nc import _parse_artemig_texto

    texto = (
        "NOTIFICAÇÃO CONSOL MG-050\n"
        "202509999 01/01/26 10:00 Parâmetros Gerais Inexistência de tachas\n"
        "LOCALIZAÇÃO\n"
        "202509999\n"
        "Rodovia: MG-050 SH16 Km Inicial: 543+500 Km Final: 544+000 Sentido: CRESCENTE\n"
        "Nº da CONSOL: 2516932\n"
    )
    nc = _parse_artemig_texto(texto)
    assert nc is not None
    assert abs(nc.km_ini - 543.5) < 1e-9
    assert abs(nc.km_fim - 544.0) < 1e-9
    assert "543" in (nc.km_ini_str or "") and "500" in (nc.km_ini_str or "")


def test_parse_artemig_texto_km_valores_antes_rotulos_tabela_pdf():
    """PyMuPDF: «653,400» nas linhas antes de «Km Inicial» / «Km Final» (BR-265)."""
    from nc_artesp.modulos.analisar_pdf_nc import _parse_artemig_texto

    texto = (
        "NOTIFICAÇÃO CONSOL\n"
        "Nº da CONSOL:\n2516932\n"
        "BR-265\nSH20\n653,400\n653,400\nCRESCENTE\nPISTA\n"
        "Rodovia\nSH\nKm Inicial\nKm Final\nSentido\nLocal\n"
        "LOCALIZAÇÃO\n202501368\n"
    )
    nc = _parse_artemig_texto(texto)
    assert nc is not None
    assert abs(nc.km_ini - 653.4) < 1e-9
    assert abs(nc.km_fim - 653.4) < 1e-9


def test_parse_artemig_texto_tabela_pipes_antes_localizacao_sinalizacao():
    """Layout com Indicador|Patologia|…|Local| antes do bloco LOCALIZAÇÃO + código na linha seguinte."""
    from nc_artesp.modulos.analisar_pdf_nc import _parse_artemig_texto

    texto = (
        "NOTIFICAÇÃO CONSOL\n"
        "Indicador|Patologia|202506787|15/09/25|10:26|"
        "Sinalização Horizontal|Placas de advertência danificadas||Local|Em até 5 dias|\n"
        "LOCALIZAÇÃO\n"
        "202506787\n"
        "MG-050 SH02 100,500 CRESCENTE PISTA\n"
        "Nº da CONSOL: 2516953\n"
    )
    nc = _parse_artemig_texto(texto)
    assert nc is not None
    assert nc.codigo == "202506787"
    assert "Sinaliza" in (nc.indicador_artemig or "")
    assert "Horizontal" in (nc.indicador_artemig or "")
    assert "Placas" in (nc.patologia_artemig or "")
    assert "Gerais (Parâmetros)" not in (nc.patologia_artemig or "")
