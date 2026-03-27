"""
modulos/nc_criar_email.py
────────────────────────────────────────────────────────────────────────────
Equivalente VBA: NC_Artesp_Criar_Email_Pad_25

A partir de e-mails selecionados no Outlook + arquivos XLS em Exportar/:

  1. Lê cada XLS individual da pasta Exportar/ (saída do mod 01).
  2. Coluna U = responsável pelo apontamento (fiscal que fez o apontamento); esse é quem recebe a resposta.
  3. Para cada planilha: identifica o e-mail selecionado cujo remetente corresponde
     ao responsável (U) e cria um único ReplyAll para esse par planilha↔fiscal.
     Assim, múltiplas planilhas (fiscais diferentes) geram respostas cada uma ao correto.
  4. Assunto/Corpo/CC como antes; salva rascunho (não envia automaticamente).

Colunas do XLS (planilha EAF individual, saída M01 – conforme macros):
  Utilizadas: C,D,E,F,H,I,J,K,L,Q,T,U,V,P. U=responsável (fiscal). P=grupo EAF. C=código fiscalização.
  Na criação do e-mail: verifica o código (e grupo/trecho) e atribui o destinatário responsável automaticamente
  via MAPA_EAF[].email (grupo da col P ou grupo obtido por rodovia+km); fallback col U se contiver @.
  Não utilizadas na extração: A,B,G,M,N,O,R,S,W,X,Y... (ver docs/ANALISE_COLUNAS_MODELO_KCOR_KRIA.md)

Igual à macro ``NC_Artesp_Criar_Email_Pad_25`` (linhas 144–145): por NC só **cabeçalho + imagem do
apontamento PDF** (711×295). As fotos **nc** e **nc_1** não são embutidas — o fiscal insere-as manualmente
no Outlook (a macro anexava-as com CID e opcionalmente punha-as no HTML; aqui replicamos só o PDF).

CIDs sequenciais ASCII (img0001…) em ``<img src="cid:…">`` e ``Content-ID: <…>``. MIME: ``related`` +
``MIMEText(html)`` + ``MIMEImage`` (fallback MIMEBase se ``email.encoders`` estiver sombreado).

Cabeçalho por linha: ``Rodovia - km H,M Sentido - Const: … - Prazo: … - Atividade - Cod. Fisc.: …``
(como ``mytext2`` na macro, antes do ``<img src=cid:pdf…>``).

Requer: Windows + Outlook instalado + pywin32
"""

import base64
import html
import logging
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook

from config import M01_EXPORTAR, M02_FOTOS_NC, M02_FOTOS_PDF, MAPA_EAF, NC_EMAIL_CC, RODOVIAS
from utils.helpers import (
    caminho_dentro_limite_windows,
    data_ddmmaaaa,
    encontrar_foto_por_codigo_ou_numero,
    escrever_bytes_caminho,
    obter_grupo_empresa_por_trecho,
    parse_data,
    path_foto_nc,
    path_foto_pdf,
    str_caminho_io_windows,
    str_caminho_outlook_mapi,
)

logger = logging.getLogger(__name__)

# Índices de coluna (1-based) – planilha EAF individual
_C  = 3   # código fiscalização
_D  = 4   # data fiscalização
_E  = 5   # horário
_F  = 6   # rodovia
_G  = 7   # concessionária
_H  = 8   # km inicial
_I  = 9   # m inicial
_J  = 10  # km final
_K  = 11  # m final
_L  = 12  # sentido
_M  = 13  # data retorno
_N  = 14  # status retorno
_O  = 15  # tipo atividade
_P  = 16  # grupo atividade
_Q  = 17  # atividade
_R  = 18  # nº notificação
_S  = 19  # data envio
_T  = 20  # data reparo
_U  = 21  # responsável
_V  = 22  # nº foto

_LINHA_INICIO = 5


def _cell(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _norm_header(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def _extrair_numero_de_pdf(nome: str) -> str:
    """
    Extrai N de 'pdf (N).jpg' / 'PDF (N).jpg'. Retorna string (mantém código se não numérico).
    """
    s = (nome or "").strip()
    if not s:
        return ""
    m = re.search(r"\(([^)]+)\)", s)
    return (m.group(1).strip() if m else s)


def _split_km_m(valor_km: str) -> tuple[str, str]:
    """
    Converte '65+800' -> ('65', '800'); '65 + 800' idem; se não tiver '+', retorna (valor, '').
    """
    s = (valor_km or "").strip().replace(" ", "")
    if "+" in s:
        a, b = s.split("+", 1)
        return a, b
    if "," in s:
        a, b = s.split(",", 1)
        return a, b
    return s, ""


def _normalizar_rodovia(valor: str) -> str:
    prefixos = {"SP 075": "SP 075", "SP 127": "SP 127",
                "SP 280": "SP 280", "SP 300": "SP 300",
                "SPI 10": "SPI 102/300"}
    for k, v in prefixos.items():
        if valor.startswith(k):
            return v
    return valor[:6].strip()


def _hdr_planilha_linha1(ws) -> dict[str, int]:
    """Mapa cabeçalho normalizado (linha 1) → índice de coluna."""
    hdr: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = _norm_header(v)
        if k and k not in hdr:
            hdr[k] = c
    return hdr


def _layout_kartado_por_hdr(hdr: dict[str, int]) -> bool:
    return ("codigo de fiscalizacao" in hdr) or ("codigo fiscalizacao" in hdr)


def _filtrar_ncs_para_email(ncs: list[dict]) -> list[dict]:
    """
    Exige código de fiscalização. Preenche atividade quando vier vazia (Kartado sem «Descrição»/«Classe»),
    para não zerar a geração de .eml como o filtro antigo if n.get('atividade').
    """
    out: list[dict] = []
    for n in ncs:
        cod = (n.get("cod") or "").strip()
        if not cod:
            continue
        if not (n.get("atividade") or "").strip():
            n["atividade"] = "Apontamento NC"
        out.append(n)
    return out


def _ler_xls(arq: Path) -> list[dict]:
    """Lê todas as NCs de um XLS/XLSX individual (saída mod 01). .xls usa a mesma conversão que o M02."""
    from utils.excel_io import xls_to_xlsx

    tmp: Path | None = None
    path_use = arq
    if arq.suffix.lower() == ".xls":
        tmp = xls_to_xlsx(arq, dest=None)
        path_use = tmp

    wb = load_workbook(str(path_use), data_only=True)
    try:
        # Tenta planilha "Sheet0" primeiro (formato VBA), depois ativa
        ws = None
        for nome in wb.sheetnames:
            if nome.lower() in ("sheet0", "sheet1", "planilha1", "folha1"):
                ws = wb[nome]
                break
        if ws is None:
            ws = wb.active

        hdr = _hdr_planilha_linha1(ws)
        if not _layout_kartado_por_hdr(hdr):
            for nome in wb.sheetnames:
                cand = wb[nome]
                h2 = _hdr_planilha_linha1(cand)
                if _layout_kartado_por_hdr(h2):
                    ws, hdr = cand, h2
                    break
        is_kartado = _layout_kartado_por_hdr(hdr)

        ncs: list[dict] = []
        if is_kartado:
            col_cod = hdr.get("codigo de fiscalizacao") or hdr.get("codigo fiscalizacao")
            col_rod = hdr.get("rodovia")
            col_km = hdr.get("km")
            col_sent = hdr.get("sentido")
            col_prazo = hdr.get("prazo")
            col_desc = hdr.get("descricao")
            col_classe = hdr.get("classe")
            col_foto2 = (
                hdr.get("foto_2")
                or hdr.get("foto 2")
                or hdr.get("foto2")
            )
            col_foto1 = (
                hdr.get("foto_1")
                or hdr.get("foto 1")
                or hdr.get("foto1")
            )
            col_encontrado = hdr.get("encontrado em")

            ultima = ws.max_row
            for r in range(ultima, 1, -1):
                if col_cod and ws.cell(row=r, column=col_cod).value:
                    ultima = r
                    break

            for r in range(2, ultima + 1):
                cod = _cell(ws, r, col_cod) if col_cod else ""
                if not cod:
                    continue
                rod_raw = _cell(ws, r, col_rod) if col_rod else ""
                km_raw = _cell(ws, r, col_km) if col_km else ""
                km_i, m_i = _split_km_m(km_raw)
                desc = _cell(ws, r, col_desc) if col_desc else ""
                if desc:
                    if "-->" in desc:
                        atividade = desc.split("-->", 1)[0].strip()
                    else:
                        atividade = desc.splitlines()[0].strip()
                else:
                    atividade = ""
                if not atividade:
                    atividade = (_cell(ws, r, col_classe) if col_classe else "")
                foto2_nome = _cell(ws, r, col_foto2) if col_foto2 else ""
                foto_ref = _extrair_numero_de_pdf(foto2_nome)
                if not foto_ref and col_foto1:
                    f1 = _cell(ws, r, col_foto1)
                    if f1:
                        foto_ref = _extrair_numero_de_pdf(f1)
                ncs.append(
                    {
                        "cod": cod,
                        "data_fisc": _cell(ws, r, col_encontrado) if col_encontrado else "",
                        "rodovia": _normalizar_rodovia((rod_raw or "").replace("-", " ")),
                        "km_i": km_i,
                        "m_i": m_i,
                        "sentido": _cell(ws, r, col_sent) if col_sent else "",
                        "atividade": atividade,
                        "data_rep": _cell(ws, r, col_prazo) if col_prazo else "",
                        "responsavel": "",
                        "foto": foto_ref,
                        "grupo": "",
                    }
                )
            return _filtrar_ncs_para_email(ncs)

        # Layout antigo (EAF individual)
        ultima = ws.max_row
        for r in range(ultima, _LINHA_INICIO - 1, -1):
            if ws.cell(row=r, column=_C).value:
                ultima = r
                break

        for r in range(_LINHA_INICIO, ultima + 1):
            rod_raw = _cell(ws, r, _F)
            ncs.append(
                {
                    "cod": _cell(ws, r, _C),
                    "data_fisc": _cell(ws, r, _D),
                    "rodovia": _normalizar_rodovia(rod_raw),
                    "km_i": _cell(ws, r, _H),
                    "m_i": _cell(ws, r, _I),
                    "sentido": _cell(ws, r, _L),
                    "atividade": _cell(ws, r, _Q),
                    "data_rep": _cell(ws, r, _T),
                    "responsavel": _cell(ws, r, _U),
                    "foto": _cell(ws, r, _V),
                    "grupo": _cell(ws, r, _P),
                }
            )
        return _filtrar_ncs_para_email(ncs)
    finally:
        wb.close()
        if tmp is not None and tmp.exists():
            tmp.unlink(missing_ok=True)


def _cid_mapi_macro(arquivo: Path) -> str:
    """
    CID igual ao das macros Outlook / VBA: espaços → %20; prefixos PDF → pdf; NC → nc.
    Ex.: «PDF (1).jpg» → pdf%20(1).jpg ; «nc (x)_1.jpg» → nc%20(x)_1.jpg
    """
    raw = arquivo.name.replace(" ", "%20")
    if len(raw) >= 3 and raw[:3].lower() == "pdf":
        return "pdf" + raw[3:]
    if len(raw) >= 2 and raw[:2].lower() == "nc":
        return "nc" + raw[2:]
    return raw.lower()


def _nome_arquivo_header_mime_seguro(path: Path) -> str:
    """
    Filenames em Content-Disposition/Location devem ser ASCII para policy.SMTP;
    caracteres fora do intervalo fazem falhar msg.as_bytes e zerar todos os .eml.
    """
    n = path.name or "image.jpg"
    t = unicodedata.normalize("NFKD", n)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"[^\x20-\x7E]+", "_", t).strip() or "image.jpg"
    return t[:180]


def _resolver_foto_pdf(pasta_fotos_pdf: Path, nc: dict) -> "tuple[Path | None, str]":
    """
    Resolve o arquivo da foto PDF para uma NC, alinhado à lógica de renomeação do extrator.
    O extrator grava por código de fiscalização (ex.: PDF (896643).jpg, PDF (NC.13.1039).jpg);
    a planilha tem col C = código e col V = número da foto.
    Retorna (path do arquivo ou None, CID para Content-ID / cid: no HTML — padrão macro).
    """
    if not pasta_fotos_pdf or not Path(pasta_fotos_pdf).is_dir():
        return (None, "")
    pasta = Path(pasta_fotos_pdf)
    cod = (nc.get("cod") or "").strip()
    foto_raw = nc.get("foto") or ""
    foto_id_txt = str(foto_raw).strip()
    numero = None
    try:
        numero = int(float(foto_id_txt))
    except (ValueError, TypeError):
        pass
    # Buscar por identificadores em ordem de confiabilidade:
    # 1) código de fiscalização da planilha; 2) identificador textual da foto; 3) número da foto.
    candidatos_codigo: list[str] = []
    if cod:
        candidatos_codigo.append(cod)
    if foto_id_txt and foto_id_txt not in candidatos_codigo:
        candidatos_codigo.append(foto_id_txt)

    for cod_cand in candidatos_codigo:
        encontrado = encontrar_foto_por_codigo_ou_numero(
            pasta, "PDF", codigo=cod_cand, numero=numero
        )
        if encontrado and encontrado.is_file():
            return (encontrado, _cid_mapi_macro(encontrado))

    # Busca combinada (mantém compatibilidade com comportamento anterior).
    encontrado = encontrar_foto_por_codigo_ou_numero(
        pasta, "PDF", codigo=cod if cod else None, numero=numero
    )
    if encontrado and encontrado.is_file():
        return (encontrado, _cid_mapi_macro(encontrado))
    # Fallback por número (compatível com PDF (1).jpg, PDF (00001).jpg)
    if numero is not None:
        p = path_foto_pdf(pasta, numero)
        if p.is_file():
            return (p, _cid_mapi_macro(p))
        p5 = path_foto_pdf(pasta, str(numero).zfill(5))
        if p5.is_file():
            return (p5, _cid_mapi_macro(p5))

    # Fallback: ZIP/servidor com subpastas (ex.: lote_*_pdfs_imagens/) — procura «PDF (…).jpg».
    cod_l = (cod or "").strip().lower()
    foto_l = foto_id_txt.lower() if foto_id_txt else ""
    for alvo in [x for x in (cod_l, foto_l) if x]:
        try:
            for f in sorted(pasta.rglob("*.jpg")):
                if not f.is_file():
                    continue
                low = f.name.lower()
                if "pdf" not in low or "(" not in low:
                    continue
                mid = _extrair_numero_de_pdf(f.name).lower()
                if mid == alvo:
                    return (f, _cid_mapi_macro(f))
        except OSError:
            pass

    # Fallback extra: algumas execuções podem gravar com prefixo diferente (ex.: "nc (...)")
    # ou com maiúsculas/minúsculas variadas. Se o identificador dentro de "(...)" casar,
    # usamos a imagem mesmo assim para não perder o anexo no e-mail.
    ids_alvo: set[str] = set()
    if cod_l:
        ids_alvo.add(cod_l)
    if foto_l:
        ids_alvo.add(foto_l)
    if numero is not None:
        ids_alvo.add(str(numero).lower())
        ids_alvo.add(str(numero).zfill(5).lower())
    if ids_alvo:
        try:
            for f in sorted(pasta.rglob("*.jpg")):
                if not f.is_file():
                    continue
                mid = _extrair_numero_de_pdf(f.name).strip().lower()
                if mid and mid in ids_alvo:
                    return (f, _cid_mapi_macro(f))
        except OSError:
            pass

    # Fallback: só nc (N).jpg / nc (código).jpg (mesma pasta de extração)
    encontrado_nc = encontrar_foto_por_codigo_ou_numero(
        pasta, "nc", codigo=cod if cod else None, numero=numero
    )
    if encontrado_nc and encontrado_nc.is_file():
        return (encontrado_nc, _cid_mapi_macro(encontrado_nc))
    if numero is not None:
        pn = path_foto_nc(pasta, numero)
        if pn.is_file():
            return (pn, _cid_mapi_macro(pn))
        pn5 = path_foto_nc(pasta, str(numero).zfill(5))
        if pn5.is_file():
            return (pn5, _cid_mapi_macro(pn5))

    # Fallback final: comparação tolerante por identificador normalizado.
    # Útil quando o nome recebeu sufixos (ex.: "_1") ou variações de formatação.
    def _norm_id(v: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", (v or "").strip().lower())

    def _digits(v: str) -> str:
        return re.sub(r"\D+", "", (v or "").strip())

    alvo_ids: set[str] = set()
    alvo_digs: set[str] = set()
    for raw in (cod, foto_id_txt, str(numero) if numero is not None else ""):
        if not raw:
            continue
        alvo_ids.add(_norm_id(raw))
        digs = _digits(raw)
        if digs:
            alvo_digs.add(digs)

    if alvo_ids or alvo_digs:
        try:
            for f in sorted(pasta.rglob("*.jpg")):
                if not f.is_file():
                    continue
                low = f.name.lower()
                if "pdf" not in low and "nc" not in low:
                    continue
                mid = _extrair_numero_de_pdf(f.name)
                mid_norm = _norm_id(mid)
                mid_digs = _digits(mid)
                if (mid_norm and mid_norm in alvo_ids) or (mid_digs and mid_digs in alvo_digs):
                    return (f, _cid_mapi_macro(f))
        except OSError:
            pass

    return (None, "")


# Dimensões alinhadas à macro NC_Artesp_Criar_Email_Pad_25 (só imagem PDF no HTML automático).
_IMG_APONT_W = 711
_IMG_APONT_H = 295


def _eh_ficheiro_pdf_apontamento(path: Path) -> bool:
    n = path.name.lower()
    return "pdf" in n and "(" in n


def _path_pdf_apontamento_para_corpo_email(
    nc: dict,
    pasta_fotos_pdf: Path | None,
) -> Path | None:
    """
    Apenas JPG de apontamento «PDF (…).jpg».
    Não usa o fallback da resolução que devolve `nc (...).jpg` (vistoria inserida manualmente).
    """
    if not pasta_fotos_pdf or not Path(pasta_fotos_pdf).is_dir():
        return None
    p, _ = _resolver_foto_pdf(Path(pasta_fotos_pdf), nc)
    if not p or not p.name:
        return None
    if not _eh_ficheiro_pdf_apontamento(p):
        return None
    try:
        if Path(str_caminho_io_windows(p)).is_file():
            return p
    except OSError:
        pass
    return None


def _cabecalho_linha_macro_vba(nc: dict) -> str:
    """
    Igual ao trecho de cabeçalho em mytext2 da macro (antes do <img>):
    Rodovia - km H,M Sentido - Const … - Prazo … - Atividade - Cod. Fisc.: …
    """
    rod = str(nc.get("rodovia") or "").strip()
    km_i = str(nc.get("km_i") or "").strip()
    m_i = str(nc.get("m_i") or "").strip()
    sent = str(nc.get("sentido") or "").strip()
    tipo = str(nc.get("atividade") or "").strip()
    cod = str(nc.get("cod") or "").strip()
    df = parse_data(nc.get("data_fisc"))
    dr = parse_data(nc.get("data_rep"))
    dc_s = data_ddmmaaaa(df) if df else str(nc.get("data_fisc") or "").strip()
    dr_s = data_ddmmaaaa(dr) if dr else str(nc.get("data_rep") or "").strip()
    return (
        f"{rod} - km {km_i},{m_i} {sent} - Const: {dc_s} - Prazo: {dr_s} - {tipo} - Cod. Fisc.: {cod}"
    )


def _cid_imagem_inline_email(seq: int) -> str:
    """
    Identificador só [a-z0-9] para Content-ID / cid: — o estilo MAPI «pdf%20(1).jpg»
    quebra vários leitores de .eml e o motor HTML do Outlook em multipart/related.
    """
    return f"img{seq:04d}"


def _img_html_cid(cid: str, w: int, h: int) -> str:
    # Sem margin:auto — isso centralizava a imagem; margin:0 alinha à esquerda como o texto.
    return (
        f'<img src="cid:{cid}" width="{w}" height="{h}" '
        'style="display:block;border:0;max-width:100%;height:auto;margin:0;" alt="">'
    )


# Estilo base alinhado ao Outlook (macro envolve a saudação em <p>).
_EMAIL_BODY_STYLE = (
    "font-family:Calibri,Arial,Helvetica,sans-serif;font-size:11pt;color:#000000;line-height:1.4;"
    "text-align:left;"
)


def _bloco_html_macro_so_pdf(nc: dict, partes_pdf: list[tuple[Path, str]]) -> str:
    """Cabeçalho como na macro + uma ou mais imagens PDF (711×295) via cid:."""
    cab = html.escape(_cabecalho_linha_macro_vba(nc), quote=False)
    inner: list[str] = [
        '<div style="margin:0 0 24px 0;padding:0 0 18px 0;border-bottom:1px solid #d0d0d0;text-align:left;">',
        f'<p style="margin:0 0 12px 0;text-align:left;"><b><u>{cab}</u></b></p>',
    ]
    for _path, cid in partes_pdf:
        inner.append(
            f'<p style="margin:12px 0 8px 0;text-align:left;">'
            f"{_img_html_cid(cid, _IMG_APONT_W, _IMG_APONT_H)}</p>"
        )
    inner.append("<BR><BR><BR><BR></div>")
    return "".join(inner)


def _bloco_html_macro_sem_pdf(nc: dict) -> str:
    """Cabeçalho macro; sem <img> PDF (fiscal cola nc / nc_1 manualmente)."""
    cab = html.escape(_cabecalho_linha_macro_vba(nc), quote=False)
    return (
        '<div style="margin:0 0 24px 0;padding:0 0 18px 0;border-bottom:1px solid #d0d0d0;text-align:left;">'
        f'<p style="margin:0 0 12px 0;text-align:left;"><b><u>{cab}</u></b></p>'
        '<p style="margin:0;color:#666;text-align:left;"><i>(Imagem PDF do apontamento não encontrada. '
        "Insira manualmente as fotos de vistoria nc e nc_1 no corpo do e-mail, se necessário.)</i></p>"
        "<BR><BR><BR><BR></div>"
    )


def _html_saudacao() -> str:
    return (
        '<p style="margin:0 0 14px 0;">Prezados,</p>'
        '<p style="margin:0 0 22px 0;">Seguem registros fotográficos das superações de não conformidade, '
        "dentro do prazo regulamentado.</p>"
    )


def _assunto_enriquecido(assunto_original: str, nc_ref: dict) -> str:
    """Monta o assunto enriquecido (ReplyAll + dados da NC)."""
    base = assunto_original.replace(" [Email Externo] ", "")
    return (
        f"{base} - {nc_ref['rodovia']} ({nc_ref['atividade']}) "
        f"- Const: {nc_ref['data_fisc']} - Prazo: {nc_ref['data_rep']}"
    )


def _cc_str() -> str:
    return ";".join(NC_EMAIL_CC)


def _email_por_grupo(grupo_val, mapa_eaf: list) -> str:
    """Retorna o e-mail do responsável do grupo EAF (MAPA_EAF[].email), se definido."""
    if not mapa_eaf:
        return ""
    if grupo_val is None or (isinstance(grupo_val, (int, float)) and grupo_val == 0):
        return ""
    try:
        g = int(float(str(grupo_val).strip()))
    except (ValueError, TypeError):
        return ""
    if g == 0:
        return ""
    for entry in mapa_eaf:
        if entry.get("grupo") == g:
            email = (entry.get("email") or "").strip()
            if email and "@" in email:
                return email
            break
    return ""


def _km_float(km_cel: str) -> float | None:
    """Converte célula de km (ex. 143+800 ou 143.8) para float."""
    if not km_cel:
        return None
    s = str(km_cel).strip().replace(",", ".")
    if "+" in s:
        parts = s.split("+", 1)
        try:
            return float(parts[0].strip()) + float(parts[1].strip()) / 1000.0
        except (ValueError, TypeError):
            return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _destinatario_responsavel_automatico(nc_ref: dict, mapa_eaf: list | None = None) -> str:
    """
    Atribui o destinatário responsável automaticamente a partir dos dados da NC (código/grupo).
    Na criação do e-mail, verifica o código (e grupo/rodovia+km) e resolve o e-mail do responsável:
    1) E-mail do grupo EAF: col P (grupo) ou grupo obtido por rodovia+km → MAPA_EAF[].email.
    2) Senão, col U (responsável) se contiver @.
    Assim o To é definido pelo grupo da fiscalização (código/trecho), sem depender só da coluna U.
    """
    mapa = mapa_eaf if mapa_eaf is not None else MAPA_EAF
    cod = nc_ref.get("cod") or ""
    email = _email_por_grupo(nc_ref.get("grupo"), mapa)
    if email:
        logger.debug("Destinatário automático (código %s, grupo col P): %s", cod, email)
        return email
    grupo, empresa = obter_grupo_empresa_por_trecho(
        nc_ref.get("rodovia"), _km_float(nc_ref.get("km_i")), mapa
    )
    email = _email_por_grupo(grupo, mapa)
    if email:
        logger.debug("Destinatário automático (código %s, grupo %s %s por trecho): %s", cod, grupo, empresa, email)
        return email
    resp = (nc_ref.get("responsavel") or "").strip()
    if resp and "@" in resp:
        return resp
    return ""


def _responsavel_casa_com_remetente(responsavel: str, sender_email: str, sender_name: str) -> bool:
    """
    Verifica se o responsável pelo apontamento (col U da planilha) corresponde ao remetente do e-mail.
    Aceita: e-mail igual, nome contido no remetente, ou e-mail contido no responsável.
    """
    r = (responsavel or "").strip().lower()
    if not r:
        return False
    se = (sender_email or "").strip().lower()
    sn = (sender_name or "").strip()
    if se and r == se:
        return True
    if sn and r in sn.lower():
        return True
    if se and se in r:
        return True
    if r and r in se:
        return True
    return False


def _criar_via_outlook(pasta_xls: Path,
                        pasta_fotos_pdf: Path,
                        pasta_fotos_nc: Path | None = None,
                        callback_progresso=None) -> int:
    """
    Para cada XLS, identifica o fiscal responsável (col U) e cria o reply no e-mail
    selecionado cujo remetente corresponda a esse fiscal. Assim cada planilha gera
    resposta para o fiscal correto (múltiplas planilhas = múltiplos fiscais).
    Retorna número de e-mails rascunhados.

    ``pasta_fotos_nc`` mantém compatibilidade da API; o corpo automático usa só ``pasta_fotos_pdf``
    (imagem PDF). Fotos nc / nc_1 são inseridas manualmente no Outlook.
    """
    _ = pasta_fotos_nc
    try:
        import win32com.client as win32
    except ImportError:
        raise ImportError(
            "pywin32 não instalado. Execute: pip install pywin32\n"
            "Requer Windows + Outlook instalado e aberto."
        )

    # Coletar XLS
    # Igual ao M02 (gerar_modelo_foto): ZIP/stage pode trazer .xlsx em subpastas de Exportar/.
    arquivos = sorted([
        f for f in pasta_xls.rglob("*.xls*")
        if f.is_file() and not f.name.startswith("~") and not f.name.startswith("_")
    ])
    if not arquivos:
        logger.warning(f"Nenhum XLS em: {pasta_xls}")
        return 0

    outlook   = win32.Dispatch("Outlook.Application")
    explorer  = outlook.ActiveExplorer()
    selection = explorer.Selection

    # Lista de e-mails selecionados com remetente (para casar com col U)
    itens_selecao = []
    for i in range(1, selection.Count + 1):
        item = selection.Item(i)
        if item.Class != 43:  # 43 = olMail
            continue
        try:
            sender_email = getattr(item, "SenderEmailAddress", "") or ""
            sender_name  = getattr(item, "SenderName", "") or ""
            itens_selecao.append((item, sender_email, sender_name))
        except Exception as e:
            logger.warning(f"  E-mail ignorado (erro ao obter remetente): {e}")
    if not itens_selecao:
        logger.warning("Nenhum e-mail de correio selecionado no Outlook.")
        return 0

    PR_ATTACH_CONTENT_ID = "http://schemas.microsoft.com/mapi/proptag/0x3712001F"
    rascunhos = 0

    for idx, arq in enumerate(arquivos):
        if callback_progresso:
            callback_progresso(idx + 1, len(arquivos), f"Processando: {arq.name[:50]}")

        ncs = _ler_xls(arq)
        if not ncs:
            logger.warning(f"  Nenhuma NC em {arq.name}")
            continue

        nc_ref = ncs[0]
        responsavel = (nc_ref.get("responsavel") or "").strip()

        # Encontrar o e-mail selecionado cujo remetente corresponde ao responsável pelo apontamento (col U)
        item_correspondente = None
        for item, sender_email, sender_name in itens_selecao:
            if _responsavel_casa_com_remetente(responsavel, sender_email, sender_name):
                item_correspondente = item
                break

        if item_correspondente is None:
            logger.warning(
                f"  Planilha {arq.name}: responsável pelo apontamento (col U) '{responsavel or '(vazio)'}' não corresponde a nenhum remetente dos e-mails selecionados. "
                "Rascunho não criado. Verifique a coluna U (responsável) e selecione os e-mails dos fiscais corretos."
            )
            continue

        reply = item_correspondente.ReplyAll()
        reply.Subject = _assunto_enriquecido(reply.Subject, nc_ref)
        reply.CC = _cc_str()
        to_addr = _destinatario_responsavel_automatico(nc_ref)
        reply.To = to_addr if to_addr else ""

        corpo_html = _html_saudacao()
        img_seq_outlook = 0
        for nc in ncs:
            pdf_path = _path_pdf_apontamento_para_corpo_email(nc, pasta_fotos_pdf)
            if not pdf_path:
                logger.warning(
                    "  Imagem PDF de apontamento não encontrada para NC cod=%s foto=%s",
                    nc.get("cod"),
                    nc.get("foto"),
                )
                corpo_html += _bloco_html_macro_sem_pdf(nc)
                continue
            img_seq_outlook += 1
            cid_u = _cid_imagem_inline_email(img_seq_outlook)
            try:
                attach = reply.Attachments.Add(str_caminho_outlook_mapi(pdf_path))
                attach.PropertyAccessor.SetProperty(PR_ATTACH_CONTENT_ID, cid_u)
            except Exception as ex_att:
                logger.warning("  Anexo Outlook omitido %s: %s", pdf_path, ex_att)
                img_seq_outlook -= 1
                corpo_html += _bloco_html_macro_sem_pdf(nc)
                continue
            corpo_html += _bloco_html_macro_so_pdf(nc, [(pdf_path, cid_u)])

        olFormatHTML = 2
        # Imagens cid: só aparecem com corpo HTML; modo texto ignora anexos inline.
        try:
            reply.BodyFormat = olFormatHTML
        except Exception as ex_bf:
            logger.warning("  BodyFormat HTML não aplicado (imagens podem falhar): %s", ex_bf)
        # Não prefixar um segundo <html><body> — o Outlook já traz um documento em HTMLBody.
        existente = reply.HTMLBody or ""
        reply.HTMLBody = f'<div style="{_EMAIL_BODY_STYLE}">{corpo_html}</div>' + existente

        reply.Save()
        rascunhos += 1
        logger.info(f"  Rascunho salvo (fiscal: {responsavel or 'N/A'}): {reply.Subject[:60]}")

    return rascunhos


def _mime_subtype_imagem(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in (".jpg", ".jpeg"):
        return "jpeg"
    if ext == ".png":
        return "png"
    if ext == ".gif":
        return "gif"
    return "jpeg"


def _parte_mime_imagem_inline(
    raw: bytes,
    sub: str,
    cid: str,
    filename: str,
):
    """
    Parte image/* inline como na documentação smtplib (MIMEImage + Content-ID + Disposition).
    Se ``MIMEImage`` falhar (ex.: ``email.encoders`` sombreado por outro pacote), usa MIMEBase + Base64.
    """
    from email.mime.base import MIMEBase

    cid = (cid or "").strip()
    try:
        import email.mime.image as mi

        part = mi.MIMEImage(raw, _subtype=sub)
    except Exception:
        part = MIMEBase("image", sub)
        part.set_payload(base64.encodebytes(raw).decode("ascii"))
        part["Content-Transfer-Encoding"] = "base64"
    part.add_header("Content-ID", f"<{cid}>")
    part.add_header("Content-Disposition", "inline", filename=filename)
    return part


def _criar_eml(pasta_xls: Path,
               pasta_fotos_pdf: Path,
               pasta_saida: Path,
               callback_progresso=None,
               pasta_fotos_nc: Path | None = None) -> list[Path]:
    """
    Gera ficheiros .eml por XLS. Mesmo padrão clássico de tutoriais smtplib: ``MIMEMultipart('related')``,
    primeiro ``MIMEText(..., 'html')``, depois imagens com ``MIMEImage`` (ou fallback MIMEBase),
    ``cid:img0001`` ↔ ``Content-ID: <img0001>``, ``as_bytes`` com ``compat32`` em primeiro lugar.

    ``pasta_fotos_nc`` mantém compatibilidade da API; só imagens PDF de apontamento entram no MIME.
    """
    _ = pasta_fotos_nc
    import email.mime.multipart as mm
    import email.mime.text      as mt
    from email import policy
    from email.header import Header
    from utils.helpers import garantir_pasta, sanitizar_nome, timestamp_agora

    garantir_pasta(pasta_saida)

    # Igual ao M02 (gerar_modelo_foto): ZIP/stage pode trazer .xlsx em subpastas de Exportar/.
    arquivos = sorted([
        f for f in pasta_xls.rglob("*.xls*")
        if f.is_file() and not f.name.startswith("~") and not f.name.startswith("_")
    ])
    gerados = []
    logger.info(
        "NC e-mail .eml: pasta_xls=%s | ficheiros *.xls*=%d | pasta_pdf=%s",
        pasta_xls,
        len(arquivos),
        pasta_fotos_pdf,
    )
    if not arquivos:
        logger.warning(
            "NC e-mail .eml: nenhum .xls/.xlsx em %s (verifique extração do ZIP para Exportar/).",
            pasta_xls,
        )

    for idx, arq in enumerate(arquivos):
        if callback_progresso:
            callback_progresso(idx + 1, len(arquivos), f"Gerando .eml: {arq.name[:50]}")
        try:
            ncs = _ler_xls(arq)
            if not ncs:
                logger.warning("  Planilha sem NC válida para e-mail: %s", arq.name)
                continue

            nc_ref = ncs[0]
            assunto = _assunto_enriquecido("RE: Apontamento NC Artesp", nc_ref)

            msg = mm.MIMEMultipart("related")
            msg["Subject"] = str(Header(assunto, "utf-8"))
            msg["From"]    = "artesp.nc@conservacao.br"
            msg["CC"]      = _cc_str()
            # Abre como rascunho não enviado no Outlook; útil para conferência antes do envio.
            msg["X-Unsent"] = "1"
            to_addr = _destinatario_responsavel_automatico(nc_ref)
            if to_addr:
                msg["To"] = to_addr

            corpo_html = _html_saudacao()
            partes_img: list[tuple[Path, str]] = []
            img_seq = 0

            for nc in ncs:
                pdf_path = _path_pdf_apontamento_para_corpo_email(nc, pasta_fotos_pdf)
                if not pdf_path:
                    logger.warning(
                        "  Imagem PDF de apontamento não encontrada para NC cod=%s foto=%s",
                        nc.get("cod"),
                        nc.get("foto"),
                    )
                    corpo_html += _bloco_html_macro_sem_pdf(nc)
                    continue
                img_seq += 1
                cid_u = _cid_imagem_inline_email(img_seq)
                partes_img.append((pdf_path, cid_u))
                corpo_html += _bloco_html_macro_so_pdf(nc, [(pdf_path, cid_u)])

            if partes_img:
                msg["X-MS-Has-Attach"] = "yes"

            # 1) HTML primeiro (como em smtplib/tutorials); 2) depois cada imagem com cid igual ao HTML.
            html_doc = (
                "<html>\n<head><meta charset=\"utf-8\" /></head>\n"
                f'<body style="{_EMAIL_BODY_STYLE}">\n'
                f'<div style="{_EMAIL_BODY_STYLE}">{corpo_html}</div>\n'
                "</body>\n</html>"
            )
            msg.attach(mt.MIMEText(html_doc, "html", "utf-8"))

            for foto_path, cid in partes_img:
                nome_hdr = _nome_arquivo_header_mime_seguro(foto_path)
                try:
                    with open(str_caminho_io_windows(foto_path), "rb") as f:
                        raw = f.read()
                except OSError as e_io:
                    logger.warning("  Ficheiro de imagem omitido do .eml: %s (%s)", foto_path, e_io)
                    continue
                sub = _mime_subtype_imagem(foto_path)
                msg.attach(_parte_mime_imagem_inline(raw, sub, cid, nome_hdr))

            stem_seguro = sanitizar_nome(arq.stem, max_len=120) or "nc"
            nome_eml = f"{timestamp_agora()} - {idx:03d} - {stem_seguro}.eml"
            destino = caminho_dentro_limite_windows(pasta_saida / nome_eml)
            payload = None
            for pol_name, pol in (("compat32", policy.compat32), ("SMTP", policy.SMTP)):
                try:
                    payload = msg.as_bytes(policy=pol)
                    break
                except Exception as ex_pol:
                    logger.warning("  as_bytes(%s) falhou para %s (%s)", pol_name, arq.name, ex_pol)
            if payload is None:
                payload = msg.as_string().encode("utf-8", errors="replace")
            escrever_bytes_caminho(destino, payload)
            gerados.append(destino)
            logger.info(f"  .eml gerado: {destino.name}")
        except Exception as e_arq:
            logger.error("  Falha ao gerar .eml para %s: %s", arq.name, e_arq, exc_info=True)
            continue

    return gerados


def executar(pasta_xls: Path | None = None,
             pasta_fotos_pdf: Path | None = None,
             pasta_fotos_nc: Path | None = None,
             usar_outlook: bool = True,
             pasta_saida_eml: Path | None = None,
             callback_progresso=None) -> dict:
    """
    Cria e-mails padrão de resposta NC.

    usar_outlook=True  → ReplyAll via Outlook COM (requer Outlook aberto
                          com e-mail(s) selecionado(s))
    usar_outlook=False → gera arquivos .eml portáteis

    Retorna dict com 'rascunhos' (int) e/ou 'eml' (list[Path]).
    """
    pasta_xls       = pasta_xls       or M01_EXPORTAR
    pasta_fotos_pdf = pasta_fotos_pdf or M02_FOTOS_PDF
    if pasta_fotos_nc is None:
        pasta_fotos_nc = M02_FOTOS_NC
    try:
        qtd_imgs = len(list(Path(pasta_fotos_pdf).rglob("*.jpg"))) if Path(pasta_fotos_pdf).is_dir() else 0
    except OSError:
        qtd_imgs = 0
    try:
        qtd_nc = len(list(Path(pasta_fotos_nc).rglob("*.jpg"))) if Path(pasta_fotos_nc).is_dir() else 0
    except OSError:
        qtd_nc = 0
    logger.info(
        "Módulo NC Email: pdf=%s (jpg=%d) | nc=%s (jpg=%d)",
        pasta_fotos_pdf,
        qtd_imgs,
        pasta_fotos_nc,
        qtd_nc,
    )

    resultado = {"rascunhos": 0, "eml": []}

    if usar_outlook:
        logger.info("Módulo NC Email: criando rascunhos via Outlook...")
        rascunhos = _criar_via_outlook(
            pasta_xls, pasta_fotos_pdf, pasta_fotos_nc, callback_progresso
        )
        resultado["rascunhos"] = rascunhos
        logger.info(f"Módulo NC Email concluído: {rascunhos} rascunho(s).")
    else:
        logger.info("Módulo NC Email: gerando arquivos .eml...")
        from config import M04_SAIDA
        pasta_eml = pasta_saida_eml or (M04_SAIDA / "emails")
        emls = _criar_eml(
            pasta_xls,
            pasta_fotos_pdf,
            pasta_eml,
            callback_progresso,
            pasta_fotos_nc=pasta_fotos_nc,
        )
        resultado["eml"] = emls
        logger.info(f"Módulo NC Email concluído: {len(emls)} arquivo(s) .eml.")

    if callback_progresso:
        callback_progresso(1, 1, "Módulo NC Email concluído.")

    return resultado
