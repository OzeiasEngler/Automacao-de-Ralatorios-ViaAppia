"""
modulos/01_separar_nc.py
────────────────────────────────────────────────────────────────────────────
Equivalente VBA: Art_011_EAF_Separar_Mod_Exc_NC
Desenvolvedor: Ozeias Engler

A partir da planilha-mãe EAF (única, com todas as NCs do período),
gera ficheiros XLS individuais. Por defeito, só junta várias linhas no mesmo Excel
se **todas** as células da linha (colunas A–última) forem iguais; data de reparo/execução
diferente implica ficheiros separados. Com `um_arquivo_por_nc=True`, uma linha por ficheiro (não consolidado).
No modo Art_011 / Template EAF, por omissão gera-se **um único** .xlsx com todas as NCs ordenadas por rodovia,
atividade e código (desativar com ``unico_arquivo_organizado=False`` ou ``um_arquivo_por_nc=True``).

Fluxo:
  Com M01_COPIA_PLANILHA_MAE=False (padrão desktop): copia o .xlsx Kartado por atividade
  (``M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO`` / ``ART03_ATIVIDADE_PARA_SERVICO_KARTADO``, ficheiros em assets/templates),
  apaga linhas de dados do template e preenche a partir da mãe. Nomes no modo Art_011 usam ``M01_SERVICO_ABREV_ART011``
  (espelho da macro Art_011); o mapa Kartado usa o texto da coluna Q da EAF, não a coluna «Classe» do template Kartado.
  Com M01_COPIA_PLANILHA_MAE=True ou ``executar(..., copia_planilha_mae=True)``: macro Art_011 — base ``Template_EAF.xlsx``
  (cabeçalho 1–4, sem dados); cada grupo recebe um ficheiro com cópia **literal** das linhas da mãe
  (valores e estilos de célula, por coluna); só a agregação por rodovia/atividade muda quais linhas vão juntas.
  A mãe continua a ser padronizada (I, K, V) antes da extração.
  API web: ``copia_planilha_mae = not m01_kartado`` (``m01_kartado=false`` → cópia mãe Art_011; ``true`` → templates Kartado).
"""

from __future__ import annotations

import logging
import re
import shutil
import unicodedata
from copy import copy
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
import xlrd

from config import (
    M01_COPIA_PLANILHA_MAE,
    M01_DICAS_PALAVRA_TEMPLATE_KARTADO,
    M01_EXPORTAR,
    M01_LINHA_INICIO,
    M01_LOTE,
    M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO,
    M01_SERVICO_ABREV_ART011,
    M01_TEMPLATE_EAF,
    PRAZO_DIAS_APOS_ENVIO,
    RODOVIA_NOME_SEPARAR,
    SERVICO_ABREV,
    SERVICO_NC,
    RODOVIAS,
)
from utils.helpers import (
    pad_metros,
    parse_data,
    data_yyyymmdd,
    km_formato_arquivo,
    normalizar_rodovia_eaf,
    garantir_pasta,
    encurtar_nome_em_pasta,
    sanitizar_nome,
    str_caminho_io_windows,
)

logger = logging.getLogger(__name__)


def _sanitizar_nome_xlsx(nome: str, max_stem: int = 380) -> str:
    """sanitizar_nome sem cortar a extensão .xlsx nem o sufixo « - Prazo - data» (limite só no stem)."""
    nome = (nome or "").strip()
    if not nome:
        return ""
    ext = Path(nome).suffix
    if ext.lower() == ".xlsx" and nome.lower().endswith(".xlsx"):
        stem = nome[: -len(ext)]
    else:
        stem, ext = Path(nome).stem, Path(nome).suffix
    # Remove duplicado estilo Windows «…~1» no stem (não entra no padrão macro das constatações).
    stem = re.sub(r"~\d+$", "", stem).rstrip(" -.")
    return sanitizar_nome(stem, max_len=max_stem) + (ext if ext else ".xlsx")


@contextmanager
def abrir_workbook(path: Path, **kwargs):
    """Context manager para garantir fechamento do workbook."""
    wb = load_workbook(str_caminho_io_windows(path), **kwargs)
    try:
        yield wb
    finally:
        wb.close()


class ValidadorArquivoEAF:
    """Valida arquivo de entrada da etapa M01."""

    EXTENSOES_VALIDAS = {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}

    @staticmethod
    def validar(arquivo: Path) -> None:
        if not arquivo.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
        if arquivo.is_dir():
            raise ValueError(
                "O caminho informado é uma pasta, não um arquivo. "
                "Selecione o arquivo Excel da planilha EAF."
            )
        if arquivo.suffix.lower() == ".pdf":
            raise ValueError(
                "O arquivo selecionado é um PDF. O passo [1/6] Separando NCs exige a PLANILHA EXCEL (EAF), não o PDF."
            )


# TEMPLATE EAF — planilha base para os arquivos gerados (cabeçalho 1–4; dados a partir da 5)
# Procura DENTRO do projeto nc_artesp (assets/Template ou assets/templates).
# Ordem: ARTESP_TEMPLATE_EAF (se definido) → nc_artesp/assets/Template → nc_artesp/assets/templates
# Aceita: Template_EAF.xlsx ou Template_EAF.xlsx.xlsx
def _caminho_template_eaf() -> Path:
    """Retorna o Path do template EAF. Procura em nc_artesp/assets/ (dentro do projeto)."""
    # 1. ARTESP_TEMPLATE_EAF (ficheiro ou pasta)
    if M01_TEMPLATE_EAF.is_file():
        return M01_TEMPLATE_EAF
    if M01_TEMPLATE_EAF.is_dir():
        for n in ("Template_EAF.xlsx", "Template_EAF.xlsx.xlsx"):
            c = M01_TEMPLATE_EAF / n
            if c.is_file():
                return c
    # 2. Pastas dentro do pacote nc_artesp (projeto GeradorARTESP)
    _nc = Path(__file__).resolve().parent.parent
    pastas = [
        _nc / "assets" / "Template",
        _nc / "assets" / "templates",
    ]
    for pasta in pastas:
        for nome in ("Template_EAF.xlsx", "Template_EAF.xlsx.xlsx"):
            candidato = pasta / nome
            if candidato.is_file():
                return candidato
    return M01_TEMPLATE_EAF  # usado na mensagem de erro se não encontrar nenhum


def _norm_stem_comparar(s: str) -> str:
    """Igual ao critério de fotos_campo.core para nomes de .xlsx (Kartado / macros)."""
    t = unicodedata.normalize("NFC", s or "")
    for u in ("\u2013", "\u2014", "\u2212"):
        t = t.replace(u, "-")
    return t.casefold()


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def _kartado_repo_template_roots() -> tuple[Path, ...]:
    """Raiz do repo: Kartado/Planilhas Padrão - Templates (e variante sem acento), se existirem."""
    kd = _repo_root() / "Kartado"
    if not kd.is_dir():
        return ()
    out: list[Path] = []
    for name in ("Planilhas Padrão - Templates", "Planilhas Padrao - Templates"):
        p = kd / name
        if p.is_dir():
            out.append(p)
    return tuple(out)


def _iter_xlsx_kartado_repo_extra() -> list[Path]:
    found: list[Path] = []
    for root in _kartado_repo_template_roots():
        try:
            for f in root.rglob("*.xlsx"):
                if _deve_excluir_xlsx_template_m01(f):
                    continue
                found.append(f)
        except OSError:
            continue
    return found


# Modelos de pipeline (Kria/Kcor/acumulado) — não usar como base do M01 ao listar/fazer match fuzzy.
_M01_EXCLUIR_NOME_XLSX_NORM: frozenset[str] = frozenset(
    _norm_stem_comparar(n)
    for n in (
        "Modelo Abertura Evento Kria Conserva Rotina.xlsx",
        "Modelo.xlsx",
        "_Planilha Modelo Kcor-Kria.XLSX",
        "Eventos Acumulado Artesp para Exportar Kria.xlsx",
    )
)


def _deve_excluir_xlsx_template_m01(f: Path) -> bool:
    if not f.is_file() or f.name.startswith("~$"):
        return True
    if "Template_EAF" in f.name:
        return True
    if "Planilha Modelo Conservação" in f.name and "Foto 2 Lados" in f.name:
        return True
    return _norm_stem_comparar(f.name) in _M01_EXCLUIR_NOME_XLSX_NORM


@lru_cache(maxsize=1024)
def _xlsx_parece_layout_kartado(path_str: str) -> bool:
    """
    Valida se o XLSX parece template/planilha Kartado (cabeçalho na linha 1).
    Evita selecionar modelos de outros fluxos (ex.: Kria/Resposta) no M01 Kartado.
    """
    p = Path(path_str)
    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
        try:
            ws = wb.active
            max_c = min(int(ws.max_column or 0), 120)
            hdr: set[str] = set()
            for c in range(1, max_c + 1):
                v = ws.cell(row=1, column=c).value
                if v is None:
                    continue
                k = _norm_header(str(v))
                if k:
                    hdr.add(k)
            # Mínimo para considerar layout Kartado válido.
            return (
                ("rodovia" in hdr)
                and ("classe" in hdr)
                and (("codigo de fiscalizacao" in hdr) or ("codigo fiscalizacao" in hdr))
            )
        finally:
            wb.close()
    except Exception:
        return False


def _iter_nc_assets_xlsx_kartado_candidatos() -> list[Path]:
    """
    .xlsx Kartado sob nc_artesp/assets: subpastas templates/Template/Kartado
    e ficheiros na raiz de assets (versionados mesmo com templates/ no .gitignore).
    """
    nc_assets = Path(__file__).resolve().parent.parent / "assets"
    out: list[Path] = []
    for sub in ("templates", "Template", "Kartado"):
        d = nc_assets / sub
        if not d.is_dir():
            continue
        try:
            for f in d.rglob("*.xlsx"):
                out.append(f)
        except OSError:
            continue
    try:
        for f in nc_assets.glob("*.xlsx"):
            out.append(f)
    except OSError:
        pass
    return out


def _resolver_ficheiro_xlsx_por_nome_em_repo(nome: str) -> Path | None:
    """Localiza `nome` em nc_artesp/assets (templates, Template, raiz), fotos_campo/assets ou repo Kartado/."""
    if not (nome or "").strip():
        return None
    alvo = _norm_stem_comparar(nome)
    try:
        for f in _iter_nc_assets_xlsx_kartado_candidatos():
            if _deve_excluir_xlsx_template_m01(f):
                continue
            if _norm_stem_comparar(f.name) != alvo:
                continue
            if _xlsx_parece_layout_kartado(str(f.resolve())):
                return f
    except OSError:
        pass
    try:
        from fotos_campo.core import (
            _ficheiro_xlsx_bundled_por_nome,
            _ficheiro_xlsx_por_nome_em_assets,
        )
    except ImportError:
        return None
    p = _ficheiro_xlsx_por_nome_em_assets(nome)
    if p is not None and p.is_file() and _xlsx_parece_layout_kartado(str(p.resolve())):
        return p
    p = _ficheiro_xlsx_bundled_por_nome(nome)
    if p is not None and p.is_file() and _xlsx_parece_layout_kartado(str(p.resolve())):
        return p
    for root in _kartado_repo_template_roots():
        try:
            for f in root.rglob("*.xlsx"):
                if _deve_excluir_xlsx_template_m01(f):
                    continue
                if _norm_stem_comparar(f.name) == alvo:
                    if _xlsx_parece_layout_kartado(str(f.resolve())):
                        return f
        except OSError:
            continue
    return None


@lru_cache(maxsize=32)
def _listar_candidatos_templates_kartado_cache() -> tuple[Path, ...]:
    """
    .xlsx em nc_artesp/assets, fotos_campo/assets (subpastas templates/Template/Kartado)
    e em repo/Kartado/Planilhas Padrão - Templates — exceto EAF, Kria/Kcor/acumulado e Foto 2 Lados.
    """
    repo = _repo_root()
    nc_assets = Path(__file__).resolve().parent.parent / "assets"
    bases = [
        nc_assets,
        repo / "fotos_campo" / "assets",
    ]
    out: list[Path] = []
    for base in bases:
        # Raiz de assets (ex.: «Dren. - Superficial - Reparo.xlsx» versionado em nc_artesp/assets/)
        try:
            for f in base.glob("*.xlsx"):
                if _deve_excluir_xlsx_template_m01(f):
                    continue
                if not _xlsx_parece_layout_kartado(str(f.resolve())):
                    continue
                out.append(f)
        except OSError:
            pass
        for sub in ("templates", "Template", "Kartado"):
            d = base / sub
            if not d.is_dir():
                continue
            try:
                for f in d.rglob("*.xlsx"):
                    if _deve_excluir_xlsx_template_m01(f):
                        continue
                    if not _xlsx_parece_layout_kartado(str(f.resolve())):
                        continue
                    out.append(f)
            except OSError:
                continue
    out.extend(_iter_xlsx_kartado_repo_extra())
    # únicos por caminho resolvido
    seen: set[str] = set()
    uniq: list[Path] = []
    for p in out:
        try:
            k = str(p.resolve())
        except OSError:
            k = str(p)
        if k not in seen:
            seen.add(k)
            uniq.append(p)
    return tuple(uniq)


def _listar_candidatos_templates_kartado() -> list[Path]:
    """Wrapper compatível que devolve list a partir do cache."""
    return list(_listar_candidatos_templates_kartado_cache())


def _tokens_atividade(s: str) -> set[str]:
    return {t for t in re.split(r"[^\w]+", (s or "").lower()) if len(t) >= 3}


def _norm_key_template_lookup(s: str) -> str:
    """
    Normaliza texto para lookup de mapa de templates:
    - remove acentos
    - remove pontuação (parênteses, ponto final, etc.)
    - colapsa espaços
    """
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.lower()
    t = re.sub(r"[^0-9a-zA-Z]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


_M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO_NORM: dict[str, str] = {
    _norm_key_template_lookup(k): v for k, v in M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO.items()
}


def _resolver_template_kartado_para_atividade(tipo_atividade: str, fallback: Path | None = None) -> Path | None:
    """
    Usado quando M01_COPIA_PLANILHA_MAE=False: escolhe o .xlsx Kartado antes de colar linhas do grupo.

    Ordem:
      1. Mapa exato M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO (texto col Q EAF, alinhado ao Art_03).
      2. Dicas por palavra-chave → ficheiro em assets.
      3. Maior sobreposição de tokens entre atividade e stem do ficheiro.
      4. fallback (opcional).
    """
    tipo = (tipo_atividade or "").strip()
    if not tipo:
        return fallback

    nome_mapa = M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO.get(tipo)
    if not nome_mapa:
        # Lookup normalizado (tolerante a pontuação/acentos/espacos finais).
        nome_mapa = _M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO_NORM.get(_norm_key_template_lookup(tipo))
    if nome_mapa:
        p = _resolver_ficheiro_xlsx_por_nome_em_repo(nome_mapa)
        if p is not None:
            return p

    candidatos = _listar_candidatos_templates_kartado()
    if not candidatos:
        return fallback

    tipo_l = tipo.lower()
    for palavra, frag in M01_DICAS_PALAVRA_TEMPLATE_KARTADO:
        if palavra not in tipo_l:
            continue
        frag_n = _norm_stem_comparar(frag)
        for f in candidatos:
            if frag_n in _norm_stem_comparar(f.stem):
                return f

    tt = _tokens_atividade(tipo)
    best: Path | None = None
    best_score = 0
    for f in candidatos:
        st = _tokens_atividade(f.stem)
        sc = len(tt & st)
        if sc > best_score:
            best_score = sc
            best = f
    if best is not None and best_score >= 1:
        return best

    # Fallback operacional: se não houver match por nome/tokens, usar qualquer template
    # Kartado válido disponível para não interromper o M01.
    if candidatos:
        return candidatos[0]

    return fallback


# ESTRUTURA — planilha-mãe EAF e Template_EAF.xlsx (fallback do M01):
#   Linhas 1 a 4 = cabeçalho; dados a partir da linha 5 (ver M01_LINHA_INICIO).
LINHA_CABECALHO_FIM = 4   # última linha do cabeçalho EAF (1–4)
PRIMEIRA_LINHA_DADOS = 5  # primeira linha de dados na mãe / Template_EAF

# Templates Kartado (nc_artesp/assets/templates, etc.): só a linha 1 é cabeçalho — não apagar; dados a partir da 2.
PRIMEIRA_LINHA_DADOS_TEMPLATE_KARTADO = 2


def _limpar_linhas_dados_eaf_no_sheet(ws, primeira_linha: int) -> None:
    """Remove linhas de dados a partir de ``primeira_linha`` (mantém cabeçalho EAF 1..primeira_linha-1)."""
    while ws.max_row >= primeira_linha:
        ws.delete_rows(ws.max_row, 1)


def _copiar_linha_mae_para_template_eaf(ws_mae, linha_mae: int, ws_tpl, linha_tpl: int, ultima_coluna: int) -> None:
    """Cópia coluna a coluna (valor + estilo) de uma linha da mãe para o template — sem remapear colunas."""
    if ultima_coluna < 1:
        return
    for col in range(1, ultima_coluna + 1):
        src = ws_mae.cell(row=linha_mae, column=col)
        dst = ws_tpl.cell(row=linha_tpl, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
    if linha_mae in ws_mae.row_dimensions and ws_mae.row_dimensions[linha_mae].height is not None:
        ws_tpl.row_dimensions[linha_tpl].height = ws_mae.row_dimensions[linha_mae].height

# CONSTANTES DE COLUNAS – planilha-mãe EAF (índice 1 = col A)
# Colunas fixas (iguais em todas as versões do EAF):
#   C=código, D=data constatação, F=rodovia, I=m_ini, K=m_fim, Q=Atividade, V=nº foto
# Coluna variável detectada dinamicamente no cabeçalho:
#   "Data Reparo" → T(20) no template manual | S(19) nos exports do sistema ARTESP
# Demais colunas (para preenchimento completo pelo módulo MA):
#   G=concessionária/EAF, H=km inicial (formato 143+800), J=km final, L=sentido, O=tipo atividade, P=grupo, U=responsável
COL_KM_I_M   = 9   # I – metros inicial
COL_KM_F_M   = 11  # K – metros final
COL_CODIGO   = 3   # C – código fiscalização / número da NC
COL_SEQ_FOTO = 22  # V – número da NC para foto (código da col C, ou sequencial)
COL_DATA_NC  = 20  # T – data reparo/prazo (fallback; detectado dinamicamente em executar())
COL_RODOVIA  = 6   # F – rodovia
COL_TIPO_NC  = 17  # Q – tipo/serviço NC (Atividade)
COL_DATA_CON = 4   # D – data da constatação
COL_CONCESSIONARIA = 7   # G – concessionária / EAF
COL_KM_I_FULL     = 8   # H – km inicial (formato 143+800)
COL_KM_F_FULL     = 10  # J – km final (formato 143+800)
COL_SENTIDO       = 12  # L – sentido
COL_TIPO_ATIV     = 15  # O – tipo atividade
COL_GRUPO_ATIV    = 16  # P – grupo (fiscalização)
COL_RESPONSAVEL   = 21  # U – responsável (fiscal)


@dataclass
class ColunasDetectadas:
    """Estrutura para colunas de datas detectadas no cabeçalho."""
    data_reparo: int
    data_envio: int | None = None


def _detectar_colunas_datas(ws, fallback_reparo: int = 20, fallback_envio: int | None = 19) -> ColunasDetectadas:
    """
    Detecta colunas de datas no cabeçalho (linhas 1-5) em uma única passada.
    - data_reparo: célula exatamente 'Data Reparo'
    - data_envio: qualquer cabeçalho contendo 'data' e 'envio'
    """
    col_reparo = None
    col_envio = None
    for r in range(1, 6):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if s == "data reparo":
                col_reparo = c
            if "data" in s and "envio" in s:
                col_envio = c
        if col_reparo is not None and col_envio is not None:
            break
    if col_reparo is None:
        col_reparo = fallback_reparo
    if col_envio is None:
        col_envio = fallback_envio
    return ColunasDetectadas(data_reparo=col_reparo, data_envio=col_envio)


def _detectar_col_data_reparo(ws, fallback: int = 20) -> int:
    """
    Detecta a coluna 'Data Reparo' lendo o cabeçalho da planilha (linhas 1-5).
    Compatível com:
      - Template manual (_Planilha Modelo nc lote 13.xls): col T(20) = 'Data Reparo'
      - Exports do sistema ARTESP: col S(19) = 'Data Reparo'
    Retorna o índice 1-based encontrado, ou `fallback` se não localizar.
    """
    cols = _detectar_colunas_datas(ws, fallback_reparo=fallback, fallback_envio=None)
    if cols.data_reparo == fallback:
        logger.warning(f"Coluna 'Data Reparo' nao encontrada no cabecalho — usando fallback col {fallback}")
    else:
        logger.debug(f"Coluna 'Data Reparo' detectada: col {cols.data_reparo}")
    return cols.data_reparo


def _detectar_col_data_envio(ws, fallback: int = 19) -> int:
    """
    Detecta a coluna 'Data do envio' ou 'Data envio' no cabeçalho (linhas 1-5).
    Usado no template EAF para mapear data da fiscalização (constatação) para a coluna correta.
    Retorna o índice 1-based; se não encontrar, retorna fallback.
    """
    cols = _detectar_colunas_datas(ws, fallback_reparo=COL_DATA_NC, fallback_envio=fallback)
    return cols.data_envio if cols.data_envio is not None else fallback


def _detectar_col_tipo_nc(ws, fallback: int = COL_TIPO_NC) -> int:
    """
    Detecta a coluna de "Atividade" no cabeçalho.
    Objetivo: retornar a coluna que contém o valor que vira template Kartado (ex.: "Defesa metálica (manutenção ou substituição)"),
    e NÃO "Tipo de Atividade" (ex.: "Segurança Rodoviária") nem colunas de data.
    """
    # Maior score = maior prioridade.
    melhor_c = None
    melhor_score = -1

    # Normaliza para evitar "Atividade", "Atividade " etc.
    def score_header(h: str) -> int:
        # Evitar pegar "Tipo de Atividade" como se fosse "Atividade".
        if "tipo" in h and "atividade" in h and "tipo de atividade" in h:
            return 1
        if "grupo" in h and "atividade" in h:
            return 1

        # Preferir explicitamente a coluna "Atividade".
        if h == "atividade":
            return 6
        if h.endswith("atividade") and "tipo" not in h and "grupo" not in h:
            return 5

        # Legado: "Evento".
        if h == "evento":
            return 3
        if "evento" in h:
            return 2

        # Outros casos possíveis (Kria/NC templates).
        if "tipo nc" in h:
            return 2
        if "servico" in h or "serviço" in h:
            return 2
        if "atividade" in h:
            # Se sobrou algo com "atividade" mas não bateu, fica abaixo.
            return 2
        return 0

    # Cabeçalho costuma estar nas primeiras linhas (1..6).
    for r in range(1, min(ws.max_row + 1, 8) + 1):
        for c in range(1, min(ws.max_column + 1, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            h = _norm_header(str(v))
            if not h:
                continue
            sc = score_header(h)
            if sc > melhor_score:
                melhor_score = sc
                melhor_c = c

    # Se não achou coluna confiável, volta ao fallback.
    return melhor_c if melhor_c is not None and melhor_score >= 3 else fallback


def _valor_tipo_nc(ws, row: int, col_tipo_nc: int):
    """Lê tipo NC evitando usar acidentalmente coluna de data."""
    v = _cell(ws, row, col_tipo_nc)
    if v and parse_data(v) is None:
        return v
    # Fallback comum no layout emergencial: coluna P (Evento)
    evento = _cell(ws, row, 16)
    if evento and parse_data(evento) is None:
        return evento
    return v


def _detectar_colunas_data_no_template(ws_template) -> tuple[int | None, int | None]:
    """
    Detecta no template EAF (cabeçalho linhas 1-5) as colunas 'Data do envio' e 'Data do reparo'.
    Retorna (col_data_envio, col_data_reparo). Se alguma não for encontrada, retorna None para essa.
    Fallback: se só 'Data Reparo' for encontrada em T(20), assume S(19) = Data do envio.
    """
    col_envio, col_reparo = None, None
    for r in range(1, 6):
        for c in range(1, min(ws_template.max_column + 1, 30)):
            v = ws_template.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if "data" in s and "envio" in s:
                col_envio = c
            if s == "data reparo":
                col_reparo = c
        if col_envio is not None and col_reparo is not None:
            break
    # Template pode ter só "Data Reparo" (ex.: col T); S(19) = Data do envio por convenção
    if col_reparo is not None and col_envio is None and col_reparo == 20:
        col_envio = 19
    return (col_envio, col_reparo)


def _converter_xls_para_xlsx(path_xls: Path) -> Path:
    """
    Lê um arquivo .xls (formato antigo) com xlrd e grava um .xlsx equivalente
    com openpyxl no mesmo diretório. Cabeçalho (linhas 1 a M01_LINHA_INICIO-1)
    e demais linhas são copiadas só com valores; não inventa cabeçalho genérico.
    Retorna o Path do arquivo .xlsx gerado.
    """
    path_xlsx = path_xls.with_suffix(".xlsx")
    if path_xlsx == path_xls:
        path_xlsx = path_xls.parent / (path_xls.stem + "_convertido.xlsx")

    book = xlrd.open_workbook(str(path_xls))
    sheet = book.sheet_by_index(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    if sheet.name:
        ws.title = sheet.name[:31]  # limite de 31 caracteres no Excel

    # Copiar só os valores do .xls (cabeçalho e dados); não gravar nada genérico.
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                continue
            excel_cell = ws.cell(row=row + 1, column=col + 1)
            if cell.ctype == xlrd.XL_CELL_NUMBER:
                excel_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                    excel_cell.value = dt
                except (ValueError, OverflowError):
                    excel_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_TEXT:
                excel_cell.value = cell.value
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                excel_cell.value = bool(cell.value)
            else:
                excel_cell.value = cell.value

    wb.save(str_caminho_io_windows(path_xlsx))
    logger.info(f"Arquivo .xls convertido para: {path_xlsx.name}")
    return path_xlsx


def _cell(ws, row: int, col: int):
    """Retorna valor da célula (row, col) ou string vazia."""
    v = ws.cell(row=row, column=col).value
    return v if v is not None else ""


def _padronizar_colunas_km(ws, row: int) -> None:
    """Padroniza colunas I e K (metros) em uma linha."""
    for col in (COL_KM_I_M, COL_KM_F_M):
        cell = ws.cell(row=row, column=col)
        cell.number_format = "@"
        cell.value = pad_metros(cell.value)


def _limpar_str(val) -> str:
    """Normaliza valor para comparação (string limpa)."""
    return str(val).strip() if val is not None else ""


def _strip_descricao_kartado_excel(s: str) -> str:
    """
    Descrição Kartado numa única linha (sem \\n na célula): remove control chars,
    quebras → espaço, colapsa espaços (evita «buracos» e spill no Excel).
    """
    if not s:
        return ""
    t = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    t = t.replace("\u2028", " ").replace("\u2029", " ")
    out: list[str] = []
    for ch in t:
        o = ord(ch)
        if ch == "\t":
            out.append(" ")
        elif o < 32:
            continue
        elif o == 0x7F:
            continue
        else:
            out.append(ch)
    return re.sub(r"\s+", " ", "".join(out)).strip()


def _kartado_data_sem_hora_celula(val) -> str | None:
    """Grava data como texto DD/MM/AAAA (sem 00:00:00) nas colunas de data do Kartado."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    dt = parse_data(val)
    if dt:
        return dt.strftime("%d/%m/%Y")
    s = str(val).strip()
    return s if s else None


def _foto_ref_numerica(val) -> str:
    """
    Referência de foto válida para nome PDF (macro usa nº/foto da linha).
    Aceita apenas numérico para evitar usar campos textuais (ex.: responsável técnico).
    """
    s = _limpar_str(val)
    if not s:
        return ""
    try:
        f = float(s.replace(",", "."))
        if f == int(f):
            return str(int(f))
    except (ValueError, TypeError):
        return ""
    return ""


def _norm_header(s: str) -> str:
    t = unicodedata.normalize("NFD", str(s or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", t).strip().lower()


def _colunas_kartado_por_header(ws) -> dict[str, int]:
    """Mapa de cabeçalho (linha 1) -> índice de coluna no template Kartado."""
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        k = _norm_header(v)
        if k and k not in out:
            out[k] = c
    return out


def _set_if_header(ws, row: int, cols: dict[str, int], header: str, valor) -> None:
    c = cols.get(_norm_header(header))
    if c:
        ws.cell(row=row, column=c).value = valor


def _fingerprint_linha_mae(
    ws, row: int, col_max: int, *, forcar_linha_unica: bool
) -> tuple[str, ...]:
    """
    Tuplo de todas as células da linha (1..col_max) normalizadas.
    Linhas só partilham o mesmo Excel se o tuplo for idêntico (incl. data reparo, km, código, etc.).
    Com forcar_linha_unica=True (um_arquivo_por_nc), acrescenta o nº da linha para nunca agrupar.
    """
    cells = tuple(_limpar_str(ws.cell(row=row, column=c).value) for c in range(1, col_max + 1))
    return cells + (str(row),) if forcar_linha_unica else cells


def _copiar_linha_com_estilo(ws_src, row_src: int, ws_dst, row_dst: int, max_col: int) -> None:
    """
    Copia uma linha inteira (valor + estilo) de ws_src para ws_dst.
    Mesma regra do gerador de modelo/foto: preserva font, border, fill, alignment, number_format.
    """
    for col in range(1, max_col + 1):
        src_cell = ws_src.cell(row=row_src, column=col)
        dst_cell = ws_dst.cell(row=row_dst, column=col)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = src_cell.font.copy()
            dst_cell.border = src_cell.border.copy()
            dst_cell.fill = src_cell.fill.copy()
            dst_cell.number_format = src_cell.number_format
            dst_cell.alignment = src_cell.alignment.copy()


def _copiar_alturas_linhas(ws_src, ws_dst, src_start: int, num_linhas: int, dst_start: int) -> None:
    """Copia as alturas das linhas do bloco origem para o bloco destino (como em gerar_modelo_foto)."""
    for offset in range(num_linhas):
        dim = ws_src.row_dimensions.get(src_start + offset)
        if dim is not None and dim.height is not None:
            ws_dst.row_dimensions[dst_start + offset].height = dim.height


def _replicar_merged_cells_header(ws_src, ws_dst, row_ini: int, row_fim: int) -> None:
    """
    Replica no sheet destino as células mescladas que estão inteiras no range [row_ini, row_fim] do origem.
    Destino usa as mesmas coordenadas (cabeçalho 1:row_fim).
    """
    for mc in list(ws_src.merged_cells.ranges):
        if mc.min_row >= row_ini and mc.max_row <= row_fim:
            try:
                ws_dst.merge_cells(
                    start_row=mc.min_row,
                    start_column=mc.min_col,
                    end_row=mc.max_row,
                    end_column=mc.max_col,
                )
            except Exception:
                pass


def atualizar_col_v_indice_global(arqs: list[Path], start_index: int) -> int:
    """
    Atualiza a coluna V (número da foto) nos XLS individuais com índice global
    sequencial (start_index+1, start_index+2, ...), na ordem dos arquivos e das linhas.
    Usado após a extração de fotos do PDF para que M02 encontre PDF (1).jpg, PDF (2).jpg, etc.
    Retorna o número de linhas atualizadas.
    """
    from openpyxl import load_workbook
    idx = start_index
    total = 0
    for path in arqs:
        if not path.exists():
            continue
        wb = load_workbook(str(path))
        ws = wb.active
        ultima = ws.max_row
        # Encontrar última linha com dados na col C
        for r in range(ultima, M01_LINHA_INICIO - 1, -1):
            if ws.cell(row=r, column=COL_CODIGO).value:
                ultima = r
                break
        for r in range(M01_LINHA_INICIO, ultima + 1):
            if ws.cell(row=r, column=COL_CODIGO).value is None:
                continue
            idx += 1
            ws.cell(row=r, column=COL_SEQ_FOTO).value = idx
            total += 1
        wb.save(str_caminho_io_windows(path))
        wb.close()
    return total


def _nome_arquivo(rodovia_raw: str, tipo_nc: str,
                  data_constatacao, data_prazo) -> str:
    """
    Monta o nome do arquivo exportado:
    yyyymmdd - CONSTATAÇÕES NC LOTE 13 (rod - serv) - Prazo - dd-mm-aaaa.xlsx
    """
    rod_info = normalizar_rodovia_eaf(rodovia_raw, RODOVIAS)
    # VBA: rod = Left(F,6); se "SPI 10" então "SPI 102-300"; senão usa rod
    rod_6    = str(rodovia_raw).strip()[:6]
    rod_nome = RODOVIA_NOME_SEPARAR.get(rod_info["tag"], rod_6) if rod_info["tag"] != "FORA" else sanitizar_nome(str(rodovia_raw)[:10])

    # Abreviação do serviço (Art_011 ElseIf serv → nome; fallback SERVICO_ABREV / sanitizar)
    tipo_st = tipo_nc.strip()
    serv_abrev = M01_SERVICO_ABREV_ART011.get(tipo_st)
    if serv_abrev:
        serv_abrev = sanitizar_nome(serv_abrev)
    else:
        serv_abrev = SERVICO_ABREV.get(tipo_st, sanitizar_nome(tipo_nc[:30]))

    # Datas: evita 00000000/00-00-0000 no nome usando data de hoje como fallback
    dt_con  = parse_data(data_constatacao)
    dt_praz = parse_data(data_prazo)
    if not dt_con:
        dt_con = datetime.now()
    if not dt_praz:
        dt_praz = datetime.now()
    yyyymmdd = data_yyyymmdd(dt_con)
    # Sufixo como na macro Art_011: " - Prazo - dd-mm-aaaa" (hífens; evita '/' no nome)
    prazo_s = dt_praz.strftime("%d-%m-%Y") if dt_praz else datetime.now().strftime("%d-%m-%Y")

    nome = (
        f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} "
        f"({rod_nome} - {serv_abrev}) - Prazo - {prazo_s}.xlsx"
    )
    return _sanitizar_nome_xlsx(nome)


def _nome_arquivo_consolidado_eaf(
    linhas_info: list,
    linhas_rows: list[int],
    linha_inicio: int,
    arquivo_mae: Path | None = None,
) -> str:
    """
    Um único Excel Art_011: data da constatação mais antiga + identificador do ficheiro-mãe
    (evita colisão ao processar vários EAF na mesma pasta).
    """
    datas_con: list[datetime] = []
    for r in linhas_rows:
        idx = r - linha_inicio
        if 0 <= idx < len(linhas_info):
            *_, data_con = linhas_info[idx]
            dt = parse_data(data_con)
            if dt:
                datas_con.append(dt)
    dt_ref = min(datas_con) if datas_con else datetime.now()
    yyyymmdd = data_yyyymmdd(dt_ref)
    stem_mae = ""
    if arquivo_mae:
        stem_mae = sanitizar_nome(Path(arquivo_mae).stem, max_len=60).strip(" -.")
    if stem_mae:
        nome = f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} - Consolidado - {stem_mae}.xlsx"
    else:
        nome = f"{yyyymmdd} - CONSTATAÇÕES NC {M01_LOTE} - Consolidado.xlsx"
    return _sanitizar_nome_xlsx(nome)


def executar(arquivo_mae: Path, pasta_destino: Path | None = None,
             callback_progresso=None, sobrescrever: bool = False,
             um_arquivo_por_nc: bool = False,
             copia_planilha_mae: bool | None = None,
             unico_arquivo_organizado: bool | None = None) -> list[Path]:
    """
    Processa a planilha-mãe EAF e gera os arquivos individuais de NC.
    sobrescrever: se True, regrava arquivos que já existem (útil em testes locais).
    um_arquivo_por_nc: se True, um Excel por linha; senão agrupa só linhas com **todas** as colunas iguais
    (incl. datas de reparo/execução — valores diferentes → ficheiros distintos). Ignorado quando a saída
    consolidada em único ficheiro está ativa (ver ``unico_arquivo_organizado``).

    copia_planilha_mae: None → usa ``M01_COPIA_PLANILHA_MAE`` (env ``ARTESP_M01_COPIA_PLANILHA_MAE``);
    True → fluxo Art_011 (``Template_EAF.xlsx`` + linhas da mãe coladas sem alterar valores); False → templates Kartado por atividade.

    unico_arquivo_organizado: None → com ``copia_planilha_mae`` True, gera **um** .xlsx com todas as NCs,
    ordenadas por rodovia, atividade e código; com Kartado (False), gera um ficheiro por atividade como antes.
    False força a separação por grupos mesmo no modo Art_011 (comportamento antigo por vários Excels).

    Com modo cópia mãe, a planilha-mãe é gravada no disco após padronizar I, K e V.

    Parâmetros
    ----------
    arquivo_mae      : Path para o .xls/.xlsx mãe (planilha EAF completa).
    pasta_destino    : Pasta onde os XLS serão salvos (padrão: M01_EXPORTAR).
    callback_progresso : função(atual, total, msg) para atualizar GUI.
    um_arquivo_por_nc : se True, um arquivo por linha da EAF (uma NC por Excel).
    copia_planilha_mae : força o modo M01 (None = configuração global).
    unico_arquivo_organizado : None = automático (único ficheiro só no modo Art_011 / Template EAF).

    Retorna
    -------
    Lista de Path dos arquivos gerados.
    """
    pasta_destino = Path(pasta_destino) if pasta_destino else M01_EXPORTAR
    garantir_pasta(pasta_destino)
    usar_copia_mae = M01_COPIA_PLANILHA_MAE if copia_planilha_mae is None else copia_planilha_mae
    if um_arquivo_por_nc:
        consolidar_um_ficheiro = False
    elif unico_arquivo_organizado is None:
        consolidar_um_ficheiro = bool(usar_copia_mae)
    else:
        consolidar_um_ficheiro = bool(unico_arquivo_organizado)
    if consolidar_um_ficheiro and not usar_copia_mae:
        raise ValueError(
            "Saída num único ficheiro organizado só é suportada com cópia planilha-mãe / Template EAF "
            "(copia_planilha_mae=True). Com templates Kartado, cada atividade gera o seu .xlsx."
        )

    ValidadorArquivoEAF.validar(arquivo_mae)
    suff = arquivo_mae.suffix.lower()

    # openpyxl só lê .xlsx/.xlsm; converter .xls quando necessário
    if suff == ".xls":
        arquivo_mae = _converter_xls_para_xlsx(arquivo_mae)
    elif suff not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        # Sem extensão ou extensão desconhecida: tentar .xls (xlrd) primeiro
        try:
            xlrd.open_workbook(str(arquivo_mae))
            arquivo_mae = _converter_xls_para_xlsx(arquivo_mae)
        except xlrd.biffh.XLRDError:
            # Pode ser .xlsx salvo sem extensão – deixar load_workbook tentar
            pass

    # Abrir só para ler: detectar colunas, última linha e lista de (data, rodovia, tipo) por linha.
    # Modo cópia mãe (Art_011): saídas usam Template_EAF.xlsx + linhas coladas da mãe (não cópia binária da mãe).
    logger.info(f"Abrindo planilha-mãe (somente leitura): {arquivo_mae.name}")
    with abrir_workbook(arquivo_mae, read_only=False) as wb_mae:
        ws = wb_mae.active

        col_data_reparo = _detectar_col_data_reparo(ws, fallback=COL_DATA_NC)
        col_tipo_nc = _detectar_col_tipo_nc(ws, fallback=COL_TIPO_NC)
        logger.info(f"Coluna 'Data Reparo': {col_data_reparo}")
        logger.info(f"Coluna 'Tipo NC/Atividade': {col_tipo_nc}")

        ultima_linha = M01_LINHA_INICIO - 1
        for r in range(ws.max_row, M01_LINHA_INICIO - 1, -1):
            if ws.cell(row=r, column=COL_CODIGO).value:
                ultima_linha = r
                break

        total_linhas = ultima_linha - M01_LINHA_INICIO + 1
        logger.info(f"Linhas de dados: {total_linhas} (L{M01_LINHA_INICIO}–L{ultima_linha})")

        linhas_info = []
        for r in range(M01_LINHA_INICIO, ultima_linha + 1):
            tipo_nc = _valor_tipo_nc(ws, r, col_tipo_nc)
            if not tipo_nc or not str(tipo_nc).strip():
                tipo_nc = _cell(ws, r, COL_CODIGO) or "NC"
            linhas_info.append((
                _cell(ws, r, col_data_reparo),
                _cell(ws, r, COL_RODOVIA),
                tipo_nc,
                _cell(ws, r, COL_DATA_CON),
            ))

        fallback_tpl = _caminho_template_eaf()
        max_col = ws.max_column

        if usar_copia_mae:
            for r in range(M01_LINHA_INICIO, ultima_linha + 1):
                _padronizar_colunas_km(ws, r)
            qseq = 1
            for r in range(M01_LINHA_INICIO, ultima_linha + 1):
                ws.cell(row=r, column=COL_SEQ_FOTO).value = qseq
                qseq += 1
            wb_mae.save(str_caminho_io_windows(arquivo_mae))
            logger.info("Planilha-mãe gravada (I, K, V) — modo cópia mãe (Art_011).")

        if consolidar_um_ficheiro and usar_copia_mae:
            candidatos: list[int] = []
            for r in range(M01_LINHA_INICIO, ultima_linha + 1):
                tipo_nc = _valor_tipo_nc(ws, r, col_tipo_nc)
                if not tipo_nc or not str(tipo_nc).strip():
                    tipo_nc = _cell(ws, r, COL_CODIGO) or "NC"
                if not tipo_nc or not str(tipo_nc).strip():
                    continue
                candidatos.append(r)

            def _chave_ordem_consolidado(rr: int) -> tuple:
                return (
                    _limpar_str(_cell(ws, rr, COL_RODOVIA)).casefold(),
                    _limpar_str(_valor_tipo_nc(ws, rr, col_tipo_nc)).casefold(),
                    _limpar_str(_cell(ws, rr, COL_CODIGO)),
                    rr,
                )

            if not candidatos:
                grupos_ord = []
            else:
                linhas_ord = sorted(candidatos, key=_chave_ordem_consolidado)
                grupos_ord = [(tuple(), linhas_ord)]
        else:
            index_fp: dict[tuple[str, ...], int] = {}
            grupos_ord = []
            for r in range(M01_LINHA_INICIO, ultima_linha + 1):
                tipo_nc = _valor_tipo_nc(ws, r, col_tipo_nc)
                if not tipo_nc or not str(tipo_nc).strip():
                    tipo_nc = _cell(ws, r, COL_CODIGO) or "NC"
                if not tipo_nc or not str(tipo_nc).strip():
                    continue
                fp = _fingerprint_linha_mae(ws, r, max_col, forcar_linha_unica=um_arquivo_por_nc)
                if fp not in index_fp:
                    index_fp[fp] = len(grupos_ord)
                    grupos_ord.append((fp, []))
                grupos_ord[index_fp[fp]][1].append(r)

    # Gerar ficheiros (reabre a mãe para ``ws`` válido durante o loop).
    with abrir_workbook(arquivo_mae, read_only=False, data_only=False) as wb_gen:
        ws = wb_gen.active
        arquivos_gerados: list[Path] = []
        processadas: set[str] = set()
        nomes_emitidos: set[str] = set()

        for fp, linhas_do_grupo in grupos_ord:
            r0 = linhas_do_grupo[0]
            idx0 = r0 - M01_LINHA_INICIO
            data_nc, rodov_raw, tipo_nc, data_con = linhas_info[idx0]

            if not tipo_nc:
                continue

            if consolidar_um_ficheiro and usar_copia_mae:
                nome_arq = _nome_arquivo_consolidado_eaf(
                    linhas_info, linhas_do_grupo, M01_LINHA_INICIO, arquivo_mae
                )
            else:
                nome_arq = _nome_arquivo(rodov_raw, tipo_nc, data_con, data_nc)
            if um_arquivo_por_nc and not consolidar_um_ficheiro:
                codigo = _cell(ws, r0, COL_CODIGO)
                codigo_safe = sanitizar_nome(str(codigo).strip())[:80] if codigo and str(codigo).strip() else f"NC-{r0}"
                stem, ext = Path(nome_arq).stem, Path(nome_arq).suffix
                nome_base = f"{stem} - {codigo_safe}{ext}"
                nome_arq = _sanitizar_nome_xlsx(nome_base)
                n = 1
                while nome_arq in processadas:
                    nome_arq = _sanitizar_nome_xlsx(f"{stem} - {codigo_safe}_{n}{ext}")
                    n += 1
                processadas.add(nome_arq)
            elif not consolidar_um_ficheiro:
                nomes_emitidos.add(nome_arq)

            destino = encurtar_nome_em_pasta(pasta_destino, nome_arq)
            garantir_pasta(pasta_destino)

            if um_arquivo_por_nc and not consolidar_um_ficheiro:
                cont = 1
                while destino.exists() and not sobrescrever:
                    stem, ext = Path(nome_arq).stem, Path(nome_arq).suffix
                    nome_arq = _sanitizar_nome_xlsx(f"{stem}_{cont}{ext}")
                    processadas.add(nome_arq)
                    destino = encurtar_nome_em_pasta(pasta_destino, nome_arq)
                    cont += 1
            elif destino.exists() and not sobrescrever:
                logger.debug("Já existe, pulando: %s", nome_arq)
                continue

            tipo_str = _limpar_str(tipo_nc)

            logger.info(
                "Gerando arquivo individual",
                extra={
                    "nome_arquivo": nome_arq,
                    "num_linhas": len(linhas_do_grupo),
                    "tipo_nc": str(tipo_str),
                    "rodovia": str(rodov_raw),
                    "modo": "copia_mae" if usar_copia_mae else "template",
                },
            )

            if usar_copia_mae:
                tpl_eaf = fallback_tpl
                if not tpl_eaf.is_file():
                    raise FileNotFoundError(
                        f"Template EAF não encontrado: {tpl_eaf}. "
                        "Coloque Template_EAF.xlsx em nc_artesp/assets/templates (ou defina ARTESP_TEMPLATE_EAF)."
                    )
                shutil.copy2(str_caminho_io_windows(tpl_eaf), str_caminho_io_windows(destino))
                with abrir_workbook(arquivo_mae, read_only=False, data_only=False) as wb_mae_linhas:
                    ws_mae_linhas = wb_mae_linhas.active
                    ultima_col = max(int(ws_mae_linhas.max_column or 0), 1)
                    with abrir_workbook(destino) as wb_out:
                        ws_out = wb_out.active
                        _limpar_linhas_dados_eaf_no_sheet(ws_out, M01_LINHA_INICIO)
                        for seq, row_orig in enumerate(sorted(linhas_do_grupo), start=0):
                            row_dest = M01_LINHA_INICIO + seq
                            _copiar_linha_mae_para_template_eaf(
                                ws_mae_linhas, row_orig, ws_out, row_dest, ultima_col
                            )
                        destino_xls = destino.with_suffix(".xls")
                        if destino_xls.exists():
                            destino_xls.unlink()
                            logger.debug("Removido .xls antigo: %s", destino_xls.name)
                        wb_out.save(str_caminho_io_windows(destino))
                arquivos_gerados.append(destino)
                logger.info("  ✓ Salvo (Template EAF + linhas da mãe): %s", destino.name)
                continue
    
            tipo_para_tpl = (_limpar_str(tipo_nc) if um_arquivo_por_nc else tipo_str) or "NC"
            template_src = _resolver_template_kartado_para_atividade(tipo_para_tpl, None)
            if template_src is None or not template_src.is_file():
                raise FileNotFoundError(
                    f"Template Kartado não encontrado para atividade '{tipo_para_tpl}'. "
                    "Verifique o .xlsx em nc_artesp/assets/ ou assets/templates/, "
                    "fotos_campo/assets/Template/, ou o mapeamento M01_MAPA_ATIVIDADE_TEMPLATE_KARTADO."
                )
            logger.info(f"  Template base: {template_src.name} (atividade={tipo_para_tpl!r})")
    
            # 1) Cópia binária do template Kartado ou fallback EAF (cabeçalho / formatação)
            shutil.copy2(str_caminho_io_windows(template_src), str_caminho_io_windows(destino))
    
            with abrir_workbook(destino) as wb_copia:
                ws_copia = wb_copia.active
                cols_tpl = _colunas_kartado_por_header(ws_copia)
    
                try:
                    mesmo_que_template_eaf = template_src.resolve() == fallback_tpl.resolve()
                except OSError:
                    mesmo_que_template_eaf = False
                primeira_linha_dados = (
                    PRIMEIRA_LINHA_DADOS if mesmo_que_template_eaf else PRIMEIRA_LINHA_DADOS_TEMPLATE_KARTADO
                )
    
                # 2) Apagar só linhas de dados: Kartado preserva cabeçalho na linha 1; Template_EAF preserva linhas 1–4
                while ws_copia.max_row >= primeira_linha_dados:
                    ws_copia.delete_rows(ws_copia.max_row, 1)
    
                # 3) Preencher linhas por cabeçalho Kartado + manter colunas técnicas para M02/M03
                for seq, row_origem in enumerate(linhas_do_grupo, start=1):
                    row_dest = primeira_linha_dados + seq - 1
                    val_envio = ws.cell(row=row_origem, column=COL_DATA_CON).value
                    val_reparo = ws.cell(row=row_origem, column=col_data_reparo).value
                    col_envio_tpl, col_reparo_tpl = _detectar_colunas_data_no_template(ws_copia)
                    if col_envio_tpl is not None:
                        ws_copia.cell(row=row_dest, column=col_envio_tpl).value = _kartado_data_sem_hora_celula(
                            val_envio
                        )
                    if col_reparo_tpl is not None:
                        if val_reparo is None or not _limpar_str(val_reparo):
                            dt_envio = parse_data(val_envio)
                            if dt_envio:
                                dt_reparo = dt_envio + timedelta(days=PRAZO_DIAS_APOS_ENVIO)
                                ws_copia.cell(row=row_dest, column=col_reparo_tpl).value = _kartado_data_sem_hora_celula(
                                    dt_reparo
                                )
                            else:
                                ws_copia.cell(row=row_dest, column=col_reparo_tpl).value = _kartado_data_sem_hora_celula(
                                    val_reparo
                                )
                        else:
                            ws_copia.cell(row=row_dest, column=col_reparo_tpl).value = _kartado_data_sem_hora_celula(
                                val_reparo
                            )
    
                    codigo = _limpar_str(ws.cell(row=row_origem, column=COL_CODIGO).value)
                    tipo_txt = _limpar_str(_valor_tipo_nc(ws, row_origem, col_tipo_nc))
                    classifica = SERVICO_NC.get(tipo_txt, ("Conservação Rotina", "Conservação Rotina", ""))[1]
                    rodovia_raw = _limpar_str(ws.cell(row=row_origem, column=COL_RODOVIA).value).replace(" ", "")
                    rodovia_fmt = {
                        "SP075": "SP-075",
                        "SP127": "SP-127",
                        "SP280": "SP-280",
                        "SP300": "SP-300",
                        "SPI102/300": "SPI-102/300",
                    }.get(rodovia_raw, rodovia_raw or "FORA")
                    km_i_int = ws.cell(row=row_origem, column=COL_KM_I_FULL).value
                    km_f_int = ws.cell(row=row_origem, column=COL_KM_F_FULL).value
                    km_i_m = ws.cell(row=row_origem, column=COL_KM_I_M).value
                    km_f_m = ws.cell(row=row_origem, column=COL_KM_F_M).value
                    # Padrão VBA: coluna "km" no Kartado recebe km+metros (ex.: '68+100').
                    km_i_formato = km_formato_arquivo(km_i_int, km_i_m)
                    km_f_formato = km_formato_arquivo(km_f_int, km_f_m)
                    sentido = ws.cell(row=row_origem, column=COL_SENTIDO).value
    
                    dt_envio = parse_data(val_envio)
                    dt_reparo = parse_data(val_reparo)
                    if dt_reparo is None and dt_envio is not None:
                        dt_reparo = dt_envio + timedelta(days=PRAZO_DIAS_APOS_ENVIO)
                    # Texto só do EAF; uma linha na célula (sem \\n — evita confusão no Excel).
                    relatorio_ref = _limpar_str(ws.cell(row=row_origem, column=COL_DATA_CON).value)
                    descricao_kartado = _strip_descricao_kartado_excel(
                        f"{tipo_txt} --> Relatório EAF Conservação Rotina nº: {relatorio_ref} "
                        f"--> Código NC: {codigo}"
                    )
    
                    foto_seq = _foto_ref_numerica(ws.cell(row=row_origem, column=COL_SEQ_FOTO).value)
                    foto_ref_nc = codigo or foto_seq
                    foto_ref_pdf = foto_seq or codigo
                    foto_1 = f"nc ({foto_ref_nc}).jpg" if foto_ref_nc else ""
                    # Mesmo padrão dos ficheiros em disco (pdf_extractor): «PDF (COD).jpg», não «pdf» minúsculo.
                    foto_2 = f"PDF ({foto_ref_pdf}).jpg" if foto_ref_pdf else ""
    
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Origem", "Artesp")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Classe", classifica or tipo_txt)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Rodovia", rodovia_fmt)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "km", km_i_formato)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "km final", km_f_formato)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Sentido", sentido)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Status", "Solicitado")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Equipe", "Sala Técnica - Soluciona")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Menu", "Não Conformidades")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Faixa", "Não se aplica")
                    _set_if_header(
                        ws_copia,
                        row_dest,
                        cols_tpl,
                        "Encontrado em",
                        _kartado_data_sem_hora_celula(dt_envio if dt_envio else val_envio),
                    )
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Código de Fiscalização", codigo)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Código Fiscalização", codigo)
                    _set_if_header(
                        ws_copia,
                        row_dest,
                        cols_tpl,
                        "Prazo",
                        _kartado_data_sem_hora_celula(dt_reparo if dt_reparo else val_reparo),
                    )
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Descrição", descricao_kartado)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Disciplina", "Conservação de Rotina")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Doc. Origem", "EAF - ROTINA")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Motivo", "Não Conformidade")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Foto_1", foto_1)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Tipo Foto_1", "Antes")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Descrição Foto_1", "Imagem - Informações Padrão")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Foto_2", foto_2)
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Tipo Foto_2", "Antes")
                    _set_if_header(ws_copia, row_dest, cols_tpl, "Descrição Foto_2", "Print PDF do Apontamento")
    
                    # Compatibilidade técnica só quando o template for EAF.
                    # Em template Kartado puro, escrever por índice fixo (C..V) corrompe colunas de negócio.
                    if mesmo_que_template_eaf:
                        ws_copia.cell(row=row_dest, column=COL_CODIGO).value = codigo
                        ws_copia.cell(row=row_dest, column=COL_DATA_CON).value = (
                            dt_envio.strftime("%d/%m/%Y") if dt_envio else val_envio
                        )
                        ws_copia.cell(row=row_dest, column=COL_RODOVIA).value = ws.cell(row=row_origem, column=COL_RODOVIA).value
                        ws_copia.cell(row=row_dest, column=COL_KM_I_FULL).value = km_i_int
                        ws_copia.cell(row=row_dest, column=COL_KM_F_FULL).value = km_f_int
                        ws_copia.cell(row=row_dest, column=COL_KM_I_M).value = _cell(ws, row_origem, COL_KM_I_M)
                        ws_copia.cell(row=row_dest, column=COL_KM_F_M).value = _cell(ws, row_origem, COL_KM_F_M)
                        _padronizar_colunas_km(ws_copia, row_dest)
                        ws_copia.cell(row=row_dest, column=COL_SENTIDO).value = sentido
                        ws_copia.cell(row=row_dest, column=COL_TIPO_NC).value = tipo_txt
                        ws_copia.cell(row=row_dest, column=COL_DATA_NC).value = (
                            dt_reparo.strftime("%d/%m/%Y") if dt_reparo else val_reparo
                        )
                        ws_copia.cell(row=row_dest, column=COL_RESPONSAVEL).value = ws.cell(row=row_origem, column=COL_RESPONSAVEL).value
                        ws_copia.cell(row=row_dest, column=COL_SEQ_FOTO).value = foto_ref_pdf
    
                destino_xls = destino.with_suffix(".xls")
                if destino_xls.exists():
                    destino_xls.unlink()
                    logger.debug(f"Removido .xls antigo: {destino_xls.name}")
    
                wb_copia.save(str_caminho_io_windows(destino))
            arquivos_gerados.append(destino)
            logger.info(f"  ✓ Salvo: {destino.name}")

    logger.info(f"Módulo 01 concluído. {len(arquivos_gerados)} arquivo(s) gerado(s).")
    if callback_progresso:
        callback_progresso(total_linhas, total_linhas, "Módulo 01 concluído.")
    return arquivos_gerados
