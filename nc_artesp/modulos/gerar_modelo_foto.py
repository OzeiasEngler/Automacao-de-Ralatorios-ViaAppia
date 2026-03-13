"""
modulos/gerar_modelo_foto.py
──────────────────────────────────────────────────────────────────────────
Equivalente VBA: Art_022_EAF_Gerar_Mod_Ft_Exc_NC
Desenvolvedor: Ozeias Engler

Para cada XLS individual gerado pelo Módulo 01 (pasta Exportar/), produz
duas saídas (sequência macros Artesp 02): Kria e Resposta.

─── SAÍDA A – Planilha Kria de Abertura de Evento ──────────────────────
  Arquivo: yyyymmdd-hhmm - {nome_sem_extensao}.xlsx
  Pasta:   Arquivos/Arquivo Foto - Conserva/
  Modelo:  Modelo Abertura Evento Kria Conserva Rotina.
  Foto:    nc (N).jpg   ← foto da vistoria de campo

─── SAÍDA B – Relatório de Resposta à Artesp ───────────────────────────
  Arquivo: yyyymmdd-hhmmss - rodoviat - dd-mm-aaaa - tipo_nc.xlsx
  Pasta:   _Respostas/_Relatório EAF - NC/Pendentes/
  Modelo:  Modelo Resposta.
  Foto:    PDF (N).jpg  ← foto extraída do PDF
"""

import logging
import warnings
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

# Suprime aviso do openpyxl: "Data Validation extension is not supported and will be removed"
warnings.filterwarnings("ignore", message=".*Data Validation.*", module="openpyxl")

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from datetime import timedelta

from utils.image_anchor import patch_add_image, get_merged_bounds
from config import (
    M01_EXPORTAR,
    M02_FOTOS_NC,
    M02_FOTOS_PDF,
    M02_MODELO_KRIA,
    M02_SALVAR_FOTO,
    M02_MODELO_RESP,
    M02_PENDENTES,
    M02_FOTO_W, M02_FOTO_H,
    M02_FOTO_PDF_W, M02_FOTO_PDF_H,
    PRAZO_DIAS_APOS_ENVIO,
    RODOVIAS,
)
from utils.excel_io import xls_to_xlsx
from utils.helpers import (
    garantir_pasta,
    parse_data,
    data_br,
    data_yyyymmdd,
    km_mais_metros,
    km_virgula_metros,
    normalizar_rodovia_eaf,
    sanitizar_nome,
    timestamp_agora,
    timestamp_completo,
    caminho_dentro_limite_windows,
)

logger = logging.getLogger(__name__)

BLOCO      = 5    # linhas por NC no Kria
BLOCO_RESP = 28   # linhas por NC no Relatório de Resposta

# Índices de coluna do XLS individual (Módulo 01)
_C  = 3   # C – código fiscalização
_D  = 4   # D – data constatação
_E  = 5   # E – horário
_F  = 6   # F – rodovia
_H  = 8   # H – km inicial (int)
_I  = 9   # I – metros inicial
_J  = 10  # J – km final
_K  = 11  # K – metros final
_L  = 12  # L – sentido
_Q  = 17  # Q – tipo NC (Atividade)
_DR = 20  # T – Data Reparo (fallback; detectado dinamicamente)
_V  = 22  # V – nº da NC para foto

_NC_NOME_ARQ = {
    "Não Conformidade": "NC",
    "Advertência":      "ADV",
    "Notificação":      "NOT",
}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS DE CAMINHO DE FOTO
# Macro: antigo nc (1).jpg; atual nc (00001).jpg (número da NC, 5 dígitos).
# ─────────────────────────────────────────────────────────────────────────────

def _codigo_estilo_ma(codigo: object) -> bool:
    """True se o código é no padrão MA (ex.: NC.13.1039, HE.13.0112). Só MA tem duas fotos por NC."""
    if codigo is None:
        return False
    s = str(codigo).strip()
    return bool(s) and "." in s and any(c.isalpha() for c in s)


def _variantes_nome_foto(prefixo: str, num: object) -> list:
    """Retorna variantes de nome: nc (1).jpg e nc (00001).jpg (5 dígitos), sem repetir."""
    out = [f"{prefixo} ({num}).jpg"]
    try:
        n = int(num)
        nome_5 = f"{prefixo} ({n:05d}).jpg"
        if nome_5 not in out:
            out.append(nome_5)
    except (TypeError, ValueError):
        pass
    return out

def path_foto_nc(pasta: Path, num: object) -> Path:
    """Foto de campo  →  nc (N).jpg ou nc (00000).jpg   — usada no relatório Kria."""
    if not pasta:
        return Path()
    if not pasta.is_dir():
        return pasta / f"nc ({num}).jpg"
    for nome in _variantes_nome_foto("nc", num):
        direto = pasta / nome
        if direto.exists():
            return direto
        for p in pasta.rglob(nome):
            if p.is_file():
                return p
    return pasta / f"nc ({num}).jpg"

def path_foto_pdf(pasta: Path, num: object) -> Path:
    """Foto extraída do PDF  →  PDF (N).jpg   — usada no relatório Resposta. Não alterar."""
    if not pasta:
        return Path()
    if not pasta.is_dir():
        return pasta / f"PDF ({num}).jpg"
    nome = f"PDF ({num}).jpg"
    direto = pasta / nome
    if direto.exists():
        return direto
    for p in pasta.rglob(nome):
        if p.is_file():
            return p
    return direto


def path_foto_nc_segunda(pasta: Path, codigo: object) -> Path:
    """Segunda foto: nc (codigo)_1.jpg. MA = col C (ex.: NC.13.1039); conservação = num (ex.: 1 ou 00001)."""
    if not pasta or codigo is None:
        return Path()
    cod = str(codigo).strip()
    if not cod:
        return Path()
    variantes = [f"nc ({cod})_1.jpg"]
    try:
        n = int(cod)
        variantes.append(f"nc ({n:05d})_1.jpg")
    except (TypeError, ValueError):
        pass
    for nome in variantes:
        direto = pasta / nome if pasta.is_dir() else pasta.parent / nome
        if direto.exists():
            return direto
        if pasta.is_dir():
            for p in pasta.rglob(nome):
                if p.is_file():
                    return p
    return Path()


# EMU: 1 cm = 914400/2.54 ≈ 360000 EMU no OOXML
_EMU_PER_CM = 914400 / 2.54
# 72 DPI para extent em px (Resposta / outros)
_EMU_PER_PX = 12700
# Tamanho ideal da foto no Kria em cm (extent = exatamente isso no Excel)
_KRIA_FOTO_W_CM = 9.70
_KRIA_FOTO_H_CM = 7.49
_PX_PER_CM = 96 / 2.54

# ─────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS DE IMAGEM
# ─────────────────────────────────────────────────────────────────────────────

def _col_px(ws, col_letter: str) -> float:
    dim = ws.column_dimensions.get(col_letter)
    width = dim.width if (dim and dim.width) else 8.0
    return width * 7 + 5


def _col_px_display(ws, col_letter: str) -> int:
    """Largura da coluna como o Excel desenha (~7 px por unidade). Usado no extent da âncora para a foto caber no merge."""
    dim = ws.column_dimensions.get(col_letter)
    width = dim.width if (dim and dim.width) else 8.0
    return max(1, int((width + 0.5) * 7))


def _row_px(ws, row: int) -> float:
    dim = ws.row_dimensions.get(row)
    height = dim.height if (dim and dim.height) else 15.0
    return height * 4 / 3


def _merged_range_px(ws, cell_addr: str):
    """Retorna (largura_px, altura_px) do merged range que contém cell_addr."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_addr)
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            w = sum(_col_px(ws, get_column_letter(c)) for c in range(mc.min_col, mc.max_col + 1))
            h = sum(_row_px(ws, r) for r in range(mc.min_row, mc.max_row + 1))
            return w, h
    col_letter = get_column_letter(col)
    return _col_px(ws, col_letter), _row_px(ws, row)


def _merged_range_px_extent(ws, cell_addr: str) -> tuple:
    """Tamanho do merge como o Excel desenha (_col_px_display), para extent da âncora — foto não ultrapassa o quadro."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_addr)
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
            w = sum(_col_px_display(ws, get_column_letter(c)) for c in range(mc.min_col, mc.max_col + 1))
            h = sum(int(_row_px(ws, r)) for r in range(mc.min_row, mc.max_row + 1))
            return max(w, 1), max(h, 1)
    col_letter = get_column_letter(col)
    return max(_col_px_display(ws, col_letter), 1), max(int(_row_px(ws, row)), 1)

def _redimensionar_imagem_bytes(img_path: Path, largura: int, altura: int) -> bytes:
    """Redimensiona para miniatura. draft() é sugestão (JPEG); resize() garante o tamanho exato; se draft falhar, segue com resize."""
    with PILImage.open(str(img_path)) as im:
        if getattr(im, "format", None) == "JPEG" and (im.width > largura or im.height > altura):
            try:
                im.draft("RGB", (int(largura), int(altura)))
            except (AttributeError, TypeError, ValueError):
                pass
        im = im.convert("RGB")
        im = im.resize((int(largura), int(altura)), PILImage.LANCZOS)
        buf = BytesIO()
        im.save(buf, format="PNG")
        return buf.getvalue()


class _ImageFromBytes(XLImage):
    """Imagem a partir de bytes em memória; openpyxl lê no save(), então mantemos referência aos bytes."""

    def __init__(self, data: bytes):
        super().__init__(BytesIO(data))
        self._bytes_data = data

    def _data(self):
        return self._bytes_data


def _tamanho_foto_kria_px() -> tuple[int, int]:
    """Pixels da imagem para 9,70 × 7,49 cm (96 DPI)."""
    w = max(1, int(_KRIA_FOTO_W_CM * _PX_PER_CM))
    h = max(1, int(_KRIA_FOTO_H_CM * _PX_PER_CM))
    return w, h


def _extent_foto_kria_emu() -> tuple[int, int]:
    """Extent da âncora em EMU para exatamente 9,70 × 7,49 cm no Excel."""
    w_emu = int(_KRIA_FOTO_W_CM * _EMU_PER_CM)
    h_emu = int(_KRIA_FOTO_H_CM * _EMU_PER_CM)
    return w_emu, h_emu


def _inserir_imagem(ws, cell_addr: str, img_path: Path, largura: int, altura: int):
    """
    Kria (célula C*): extent em cm = 9,70 × 7,49 cm. Resposta (B*): usa largura×altura em px.
    """
    from openpyxl.utils import coordinate_to_tuple

    row_num, col_num = coordinate_to_tuple(cell_addr)
    min_col, min_row, max_col, max_row = get_merged_bounds(ws, col_num, row_num)
    if cell_addr.startswith("C"):
        w, h = _tamanho_foto_kria_px()
        w_emu, h_emu = _extent_foto_kria_emu()
    else:
        w, h = max(1, largura), max(1, altura)
        w_emu = int(w * _EMU_PER_PX)
        h_emu = int(h * _EMU_PER_PX)

    data = _redimensionar_imagem_bytes(img_path, w, h)
    xl_img = _ImageFromBytes(data)
    xl_img.width = w
    xl_img.height = h

    anchor = OneCellAnchor()
    anchor._from = AnchorMarker(col=min_col - 1, colOff=0, row=min_row - 1, rowOff=0)
    anchor.ext = XDRPositiveSize2D(cx=w_emu, cy=h_emu)
    xl_img.anchor = anchor
    ws.add_image(xl_img)


def _inserir_duas_imagens_ma(
    ws, cell_addr: str, path1: Path, path2: Path, largura: int, altura: int
) -> None:
    """
    Meio Ambiente: insere duas fotos NC no mesmo quadro do Kria.
    Tamanho fixo ideal 9,70 cm × 7,49 cm. Primeira no canto superior esquerdo, segunda no inferior direito.
    """
    from openpyxl.utils import coordinate_to_tuple

    w_px, h_px = _tamanho_foto_kria_px()
    w_emu, h_emu = _extent_foto_kria_emu()
    row, col = coordinate_to_tuple(cell_addr)
    min_col, min_row, max_col, max_row = get_merged_bounds(ws, col, row)
    half_w = w_emu // 2
    half_h = h_emu // 2
    half_w_px = max(1, w_px // 2)
    half_h_px = max(1, h_px // 2)

    for idx, (img_path, col_off_emu, row_off_emu, ext_w_emu, ext_h_emu, ext_w_px, ext_h_px) in enumerate([
        (path1, 0, 0, half_w, half_h, half_w_px, half_h_px),
        (path2, half_w, half_h, half_w, half_h, half_w_px, half_h_px),
    ]):
        if not img_path or not img_path.is_file():
            continue
        data = _redimensionar_imagem_bytes(img_path, ext_w_px, ext_h_px)
        xl_img = _ImageFromBytes(data)
        xl_img.width = ext_w_px
        xl_img.height = ext_h_px
        anchor = OneCellAnchor()
        anchor._from = AnchorMarker(
            col=min_col - 1, row=min_row - 1,
            colOff=col_off_emu, rowOff=row_off_emu,
        )
        anchor.ext = XDRPositiveSize2D(cx=ext_w_emu, cy=ext_h_emu)
        xl_img.anchor = anchor
        ws.add_image(xl_img)

def _copiar_alturas_linhas(ws, src_start: int, num_linhas: int, dst_start: int):
    for offset in range(num_linhas):
        dim = ws.row_dimensions.get(src_start + offset)
        if dim is not None and dim.height is not None:
            ws.row_dimensions[dst_start + offset].height = dim.height

def _replicar_merged_cells(ws, row_ini_src: int, row_fim_src: int, row_ini_dst: int):
    desloc = row_ini_dst - row_ini_src
    ranges_a_replicar = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= row_ini_src and mc.max_row <= row_fim_src:
            ranges_a_replicar.append((mc.min_row, mc.max_row, mc.min_col, mc.max_col))
    for min_r, max_r, min_c, max_c in ranges_a_replicar:
        try:
            ws.merge_cells(
                start_row=min_r + desloc, start_column=min_c,
                end_row=max_r + desloc,   end_column=max_c,
            )
        except Exception:
            pass

# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DAS NCs
# ─────────────────────────────────────────────────────────────────────────────

def _detectar_col_data_reparo(ws, fallback: int = _DR) -> int:
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col).value
        if val and "reparo" in str(val).lower():
            return col
    return fallback

def _cell(ws, row, col):
    return ws.cell(row=row, column=col).value

def _detectar_linha_inicio_dados(ws, max_busca: int = 15) -> int:
    """Detecta a primeira linha com código na coluna C (não é cabeçalho). Fallback 5."""
    for r in range(1, min(max_busca + 1, ws.max_row + 1)):
        val = _cell(ws, r, _C)
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue
        # Pular linhas de cabeçalho
        if s.upper().startswith("CÓDIGO") or "fiscalização" in s.lower() or s.lower() == "cod":
            continue
        return r
    return 5


def _ler_ncs(ws, linha_inicio: int = 5) -> list:
    linha_inicio = _detectar_linha_inicio_dados(ws) if linha_inicio == 5 else linha_inicio
    col_data_reparo = _detectar_col_data_reparo(ws, fallback=_DR)
    ultima = ws.max_row
    for r in range(ultima, linha_inicio - 1, -1):
        if ws.cell(row=r, column=_D).value:
            ultima = r
            break

    ncs = []
    for r in range(linha_inicio, ultima + 1):
        cod      = str(_cell(ws, r, _C) or "").strip()
        data_con = parse_data(_cell(ws, r, _D))
        hora     = str(_cell(ws, r, _E) or "").strip()
        rod_raw  = str(_cell(ws, r, _F) or "").strip()
        km_i_int = _cell(ws, r, _H)
        km_i_met = str(_cell(ws, r, _I) or "").strip()
        km_f_int = _cell(ws, r, _J)
        km_f_met = str(_cell(ws, r, _K) or "").strip()
        sentido  = str(_cell(ws, r, _L) or "").strip()
        tipo_nc  = str(_cell(ws, r, _Q) or "").strip()
        num_foto = _cell(ws, r, _V)
        data_rep = parse_data(_cell(ws, r, col_data_reparo))

        if not cod:
            continue

        rod_info = normalizar_rodovia_eaf(rod_raw, RODOVIAS)
        try:
            num_foto_val = int(num_foto) if num_foto is not None and str(num_foto).strip() else 0
        except (TypeError, ValueError):
            num_foto_val = 0
        # MA: col C = mesmo código da fiscalização que renomeia a foto (nunca "Lote").
        # Conservação: col V (num_foto). Usar col C quando preenchida e não for Lote.
        cod_ok = (cod and str(cod).strip() and not str(cod).strip().upper().startswith("LOTE"))
        foto_id = cod if cod_ok else num_foto_val
        # Prazo em dias = data vencimento (reparo) − data constatação (para coluna L do Kria)
        prazo_dias = None
        if data_con is not None and data_rep is not None:
            try:
                prazo_dias = (data_rep - data_con).days
            except (TypeError, AttributeError):
                pass

        ncs.append({
            "codigo":      cod,
            "data_con":    data_con,
            "hora":        hora,
            "rod_raw":     rod_raw,
            "rod_codigo":  rod_info["codigo"],
            "rod_tag":     rod_info["tag"],
            "rod_n":       rod_info["n"],
            "km_i":        km_mais_metros(km_i_int, km_i_met),
            "km_i_virg":   km_virgula_metros(km_i_int, km_i_met),
            "km_f":        km_mais_metros(km_f_int, km_f_met),
            "km_f_virg":   km_virgula_metros(km_f_int, km_f_met),
            "sentido":     sentido,
            "tipo_nc":     tipo_nc,
            "num_foto":    num_foto_val,
            "foto_id":     foto_id,
            "prazo_dias":  prazo_dias,
            "data_reparo": data_rep,
        })

    return ncs

# ─────────────────────────────────────────────────────────────────────────────
# SAÍDA A – Planilha Kria  →  foto: nc (N).jpg
# ─────────────────────────────────────────────────────────────────────────────

def _gerar_kria(
    ncs: list, nome_base: str,
    modelo: "Path | bytes", pasta_saida: Path,
    pasta_fotos_nc: Path,          # ← fonte principal: nc (N).jpg
    relatorio: str,
    pasta_fotos_pdf_fallback: "Path | None" = None,  # ← fallback (ZIP Extrair PDF tem nc + PDF na mesma pasta)
) -> "Path | None":
    """
    Preenche o modelo Kria com as NCs lidas e insere a foto de campo.
    modelo: Path do .xlsx ou bytes já carregados (reutilização mais rápida).

    Regra de foto (SAÍDA A):
        Usa nc (N).jpg em pasta_fotos_nc; se não existir, tenta pasta_fotos_pdf_fallback
        (ZIP do Extrair PDF coloca nc (CODIGO).jpg e PDF (CODIGO).jpg na mesma pasta).

    Estrutura do bloco (5 linhas, âncora j=8 para a 1ª NC):
      j-2: B=seq,  C=tipo_nc
      j-1: C=âncora foto nc (275×210),  G=data_envio (data constatação) — M03 lê como embasamento e grava na col Data Envio do Kcor-Kria
      j:   D=rodovia,  F=sentido,  G=tipo_nc
      j+1: D=km_i,  F=km_f,  H=codigo,  L=num_foto
      j+2: C="Vencimento",  D=data_reparo,  F=data_con,  H=relatorio,  L=prazo
    """
    if isinstance(modelo, Path):
        if not modelo.exists():
            logger.error(f"Modelo Kria não encontrado: {modelo}")
            return None
        modelo_src = str(modelo)
    else:
        modelo_src = BytesIO(modelo)

    garantir_pasta(pasta_saida)
    nome_arq = f"{timestamp_agora()} - {nome_base}.xlsx"
    destino  = caminho_dentro_limite_windows(pasta_saida / nome_arq)
    garantir_pasta(destino.parent)

    wb = load_workbook(modelo_src)
    ws = wb.active
    patch_add_image(ws)

    n_ncs = len(ncs)
    if n_ncs == 0:
        logger.warning("Nenhuma NC para inserir na planilha Kria.")
        wb.save(str(destino))
        wb.close()
        return destino

    J_INICIO  = 8               # âncora da 1ª NC
    SRC_START = J_INICIO - 2   # linha 6 = início do bloco modelo

    # Expandir blocos extras para NC 2 em diante
    for extra in range(1, n_ncs):
        dst_start = SRC_START + extra * BLOCO
        ws.insert_rows(dst_start, BLOCO)
        for offset in range(BLOCO):
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=SRC_START + offset, column=col)
                dst_cell = ws.cell(row=dst_start  + offset, column=col)
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font          = copy(src_cell.font)
                    dst_cell.border        = copy(src_cell.border)
                    dst_cell.fill          = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment     = copy(src_cell.alignment)
        _copiar_alturas_linhas(ws, SRC_START, BLOCO, dst_start)
        _replicar_merged_cells(ws, SRC_START, SRC_START + BLOCO - 1, dst_start)

    # Preencher cada NC em ordem direta (NC[0] = topo)
    for idx, nc in enumerate(ncs):
        j  = J_INICIO + idx * BLOCO
        dr = nc["data_reparo"]
        dc = nc["data_con"]
        # Data do reparo = data do envio + 10 dias quando não informada
        if not dr and dc:
            dr = dc + timedelta(days=PRAZO_DIAS_APOS_ENVIO)

        # Coluna L (12) linha j+2 (ex.: 10): dias de prazo = vencimento − data constatação (ex.: "7 dias")
        prazo_dias = nc.get("prazo_dias")
        if prazo_dias is None and dr is not None and dc is not None:
            try:
                prazo_dias = (dr - dc).days
            except (TypeError, AttributeError):
                pass
        prazo_l = f"{prazo_dias} dias" if prazo_dias is not None else ""

        ws.cell(row=j - 2, column=2).value  = idx + 1
        ws.cell(row=j - 2, column=3).value  = nc["tipo_nc"]
        ws.cell(row=j - 1, column=7).value  = data_br(dc) if dc else ""  # G = data envio (constatação); M03 usa para col Data Envio do Kcor-Kria
        ws.cell(row=j,     column=4).value  = nc["rod_codigo"]
        ws.cell(row=j,     column=6).value  = nc["sentido"]
        ws.cell(row=j,     column=7).value  = nc["tipo_nc"]
        ws.cell(row=j + 1, column=4).value  = nc["km_i"]
        ws.cell(row=j + 1, column=6).value  = nc["km_f"]
        ws.cell(row=j + 1, column=8).value  = nc["codigo"]
        ws.cell(row=j + 1, column=12).value = nc.get("foto_id", nc["num_foto"])
        ws.cell(row=j + 2, column=3).value  = "Vencimento"
        ws.cell(row=j + 2, column=4).value  = data_br(dr) if dr else ""
        ws.cell(row=j + 2, column=6).value  = data_br(dc) if dc else ""
        ws.cell(row=j + 2, column=8).value  = relatorio
        ws.cell(row=j + 2, column=12).value = prazo_l

        # ── Foto: col C = Código da Fiscalização (ex.: 902531) identifica todas as fotos; fallback col V ──
        cell_foto = f"C{j - 1}"
        foto_id = nc.get("foto_id", nc["num_foto"])
        foto_path = path_foto_nc(pasta_fotos_nc, foto_id)
        if not foto_path.is_file() and pasta_fotos_pdf_fallback and pasta_fotos_pdf_fallback.is_dir():
            foto_path = path_foto_nc(pasta_fotos_pdf_fallback, foto_id)
        foto2_path = path_foto_nc_segunda(pasta_fotos_nc, foto_id)
        if not foto2_path.is_file() and pasta_fotos_pdf_fallback and pasta_fotos_pdf_fallback.is_dir():
            foto2_path = path_foto_nc_segunda(pasta_fotos_pdf_fallback, foto_id)
        if foto_path.is_file() and foto2_path.is_file():
            _inserir_duas_imagens_ma(ws, cell_foto, foto_path, foto2_path, M02_FOTO_W, M02_FOTO_H)
            logger.debug(f"  [Kria] Duas fotos nc inseridas (MA): {foto_path.name} + {foto2_path.name} → {cell_foto}")
        elif foto_path.is_file():
            _inserir_imagem(ws, cell_foto, foto_path, M02_FOTO_W, M02_FOTO_H)
            logger.debug(f"  [Kria] Foto nc inserida: {foto_path.name} → {cell_foto}")
        else:
            logger.warning(f"  [Kria] Foto não encontrada: nc ({foto_id}).jpg")

    wb.save(str(destino))
    wb.close()
    logger.info(f"Saída A (Kria) salva: {destino.name}")
    return destino


# ─────────────────────────────────────────────────────────────────────────────
# SAÍDA B – Relatório de Resposta à Artesp  →  foto: PDF (N).jpg
# ─────────────────────────────────────────────────────────────────────────────

def _gerar_resposta(
    ncs: list, modelo: "Path | bytes",
    pasta_saida: Path,
    pasta_fotos_pdf: Path,         # ← única fonte de foto para a Resposta
) -> "Path | None":
    """
    Preenche o modelo de resposta (28 linhas por NC) com cabeçalho e fotos.
    modelo: Path do .xlsx ou bytes já carregados (reutilização mais rápida).

    Regra de foto (SAÍDA B):
        Usa EXCLUSIVAMENTE  PDF (N).jpg  da pasta_fotos_pdf.
        NÃO utiliza foto nc aqui.

    Linha 1  (B1): cabeçalho longo da 1ª NC
    Linha 2  (B2): cabeçalho curto / âncora da foto PDF (480×202)
    NC 2 em diante: duplica bloco 1→28 ABAIXO do anterior e repete o padrão.
    """
    if isinstance(modelo, Path):
        if not modelo.exists():
            logger.error(f"Modelo de resposta não encontrado: {modelo}")
            return None
        modelo_src = str(modelo)
    else:
        modelo_src = BytesIO(modelo)
    if not ncs:
        return None

    garantir_pasta(pasta_saida)

    nc1      = ncs[0]
    dr       = nc1["data_reparo"]
    dc       = nc1["data_con"]
    tipo_arq = _NC_NOME_ARQ.get(nc1["tipo_nc"], sanitizar_nome(nc1["tipo_nc"]))

    dr_str = dr.strftime("%d-%m-%Y") if dr else "00-00-0000"
    dc_str = dc.strftime("%d-%m-%Y") if dc else "00-00-0000"
    dr_br  = data_br(dr) if dr else ""
    dc_br  = data_br(dc) if dc else ""

    data_para_nome = dr or dc
    if not data_para_nome:
        for nc in ncs:
            data_para_nome = nc.get("data_reparo") or nc.get("data_con")
            if data_para_nome:
                break
    if not data_para_nome:
        data_para_nome = datetime.today()

    nome_arq = (
        f"{timestamp_completo()} - {nc1['rod_tag']} - "
        f"{dr_str} - {tipo_arq}.xlsx"
    )
    destino = caminho_dentro_limite_windows(pasta_saida / nome_arq)

    wb = load_workbook(modelo_src)
    ws = wb.active
    patch_add_image(ws)

    SRC_START = 1  # bloco modelo começa na linha 1

    def _cabecalho_curto(nc: dict, dr_: str, dc_: str) -> str:
        return (
            f"{dc_} - {nc['rod_tag']} - {nc['km_i']} - "
            f"{nc['sentido']} - {dr_} - {nc['tipo_nc']} - {nc['codigo']}"
        )

    # ── NC 1: B1 vazio; B2 texto (atrás da imagem) ────────────────────────────
    ws.cell(row=1, column=2).value = ""
    ws.cell(row=2, column=2).value = _cabecalho_curto(nc1, dr_str, dc_str)

    # Foto PDF: MA = PDF (codigo).jpg; conservação = PDF (num_foto).jpg
    foto_id1 = nc1.get("foto_id", nc1["num_foto"])
    foto1 = path_foto_pdf(pasta_fotos_pdf, foto_id1)
    if foto1.exists():
        _inserir_imagem(ws, "B2", foto1, M02_FOTO_PDF_W, M02_FOTO_PDF_H)
        logger.debug(f"  [Resposta] Foto PDF inserida: {foto1.name} → B2")
    else:
        logger.warning(f"  [Resposta] Foto PDF não encontrada: PDF ({foto_id1}).jpg")

    # ── NCs 2..N: duplicar bloco ABAIXO do anterior ───────────────────────────
    linha = 1  # ponteiro para o início do bloco atual

    for nc in ncs[1:]:
        dst_start = linha + BLOCO_RESP  # sempre abaixo → ordem direta

        ws.insert_rows(dst_start, BLOCO_RESP)

        for offset in range(BLOCO_RESP):
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=SRC_START + offset, column=col)
                dst_cell = ws.cell(row=dst_start  + offset, column=col)
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font          = copy(src_cell.font)
                    dst_cell.border        = copy(src_cell.border)
                    dst_cell.fill          = copy(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment     = copy(src_cell.alignment)

        _copiar_alturas_linhas(ws, SRC_START, BLOCO_RESP, dst_start)
        _replicar_merged_cells(ws, SRC_START, SRC_START + BLOCO_RESP - 1, dst_start)

        nc_dr    = nc["data_reparo"]
        nc_dc    = nc["data_con"]
        nc_dr_s  = nc_dr.strftime("%d-%m-%Y") if nc_dr else "00-00-0000"
        nc_dc_s  = nc_dc.strftime("%d-%m-%Y") if nc_dc else "00-00-0000"
        # B(dst_start) vazio; B(dst_start+1) texto (atrás da imagem)
        ws.cell(row=dst_start,     column=2).value = ""
        ws.cell(row=dst_start + 1, column=2).value = _cabecalho_curto(nc, nc_dr_s, nc_dc_s)

        # Foto PDF: MA = PDF (codigo).jpg; conservação = PDF (num_foto).jpg
        foto_id_v = nc.get("foto_id", nc["num_foto"])
        foto_v    = path_foto_pdf(pasta_fotos_pdf, foto_id_v)
        cell_foto = f"B{dst_start + 1}"
        if foto_v.exists():
            _inserir_imagem(ws, cell_foto, foto_v, M02_FOTO_PDF_W, M02_FOTO_PDF_H)
            logger.debug(f"  [Resposta] Foto PDF inserida: {foto_v.name} → {cell_foto}")
        else:
            logger.warning(f"  [Resposta] Foto PDF não encontrada: PDF ({foto_id_v}).jpg")

        linha = dst_start  # avança ponteiro

    wb.save(str(destino))
    wb.close()
    logger.info(f"Saída B (Resposta) salva: {destino.name}")
    return destino


# ─────────────────────────────────────────────────────────────────────────────
# FUNÇÃO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def executar(
    pasta_xls: "Path | None" = None,
    modelo_kria: "Path | None" = None,
    pasta_saida_kria: "Path | None" = None,
    modelo_resposta: "Path | None" = None,
    pasta_saida_resp: "Path | None" = None,
    pasta_fotos_nc: "Path | None" = None,
    pasta_fotos_pdf: "Path | None" = None,
    callback_progresso=None,
) -> dict:
    """
    Processa todos os XLS da pasta de entrada e gera:
      - Kria    (Saída A) com fotos  nc (N).jpg
      - Resposta (Saída B) com fotos  PDF (N).jpg

    Retorna dict: { 'kria': [...], 'resposta': [...], 'erros': [...] }
    """
    pasta_xls        = pasta_xls        or M01_EXPORTAR
    modelo_kria      = modelo_kria      or M02_MODELO_KRIA
    pasta_saida_kria = pasta_saida_kria or M02_SALVAR_FOTO
    modelo_resposta  = modelo_resposta  or M02_MODELO_RESP
    pasta_saida_resp = pasta_saida_resp or M02_PENDENTES
    pasta_fotos_nc   = pasta_fotos_nc   or M02_FOTOS_NC    # nc (N).jpg
    pasta_fotos_pdf  = pasta_fotos_pdf  or M02_FOTOS_PDF   # PDF (N).jpg

    garantir_pasta(pasta_saida_kria)
    garantir_pasta(pasta_saida_resp)

    # Buscar em pasta_xls e em subpastas (ZIP pode ter Exportar/arquivo.xlsx)
    arquivos = sorted([
        f for f in pasta_xls.rglob("*.xls*")
        if f.is_file() and not f.name.startswith("~") and not f.name.startswith("_")
    ])

    if not arquivos:
        logger.warning(f"Nenhum XLS encontrado em: {pasta_xls}")
        return {"kria": [], "resposta": [], "erros": []}

    logger.info(f"Módulo 02: {len(arquivos)} arquivo(s) para processar.")
    resultados = {"kria": [], "resposta": [], "erros": []}

    # Carregar modelos uma vez em memória (evita N leituras em disco — acelera muito)
    modelo_kria_bytes = None
    modelo_resp_bytes = None
    if modelo_kria and Path(modelo_kria).exists():
        modelo_kria_bytes = Path(modelo_kria).read_bytes()
    if modelo_resposta and Path(modelo_resposta).exists():
        modelo_resp_bytes = Path(modelo_resposta).read_bytes()

    for idx, arq in enumerate(arquivos):
        if callback_progresso:
            callback_progresso(idx + 1, len(arquivos), f"Processando: {arq.name[:60]}")

        logger.info(f"\n── Processando [{idx+1}/{len(arquivos)}]: {arq.name}")

        try:
            if arq.suffix.lower() == ".xls":
                path_para_abrir = xls_to_xlsx(arq, dest=None)
                remover_depois  = True
            else:
                path_para_abrir = arq
                remover_depois  = False

            wb = load_workbook(str(path_para_abrir), data_only=True)
            try:
                ws  = wb.active
                ncs = _ler_ncs(ws)
            finally:
                wb.close()
                if remover_depois and path_para_abrir.exists():
                    path_para_abrir.unlink(missing_ok=True)

            if not ncs:
                logger.warning(f"  Nenhuma NC encontrada em {arq.name}, pulando.")
                continue

            logger.info(f"  {len(ncs)} NC(s) lida(s).")

            nome_base = arq.stem
            relatorio = nome_base[:8]

            # ── Saída A: Kria  →  nc (N).jpg ─────────────────────────────────
            modelo_kria_eff = modelo_kria_bytes if modelo_kria_bytes is not None else modelo_kria
            arq_kria = _gerar_kria(
                ncs, nome_base,
                modelo_kria_eff, pasta_saida_kria,
                pasta_fotos_nc,
                relatorio,
                pasta_fotos_pdf_fallback=pasta_fotos_pdf,  # ZIP Extrair PDF tem nc+PDF na mesma pasta
            )
            if arq_kria:
                resultados["kria"].append(arq_kria)

            # ── Saída B: Resposta  →  PDF (N).jpg ────────────────────────────
            modelo_resp_eff = modelo_resp_bytes if modelo_resp_bytes is not None else modelo_resposta
            arq_resp = _gerar_resposta(
                ncs,
                modelo_resp_eff, pasta_saida_resp,
                pasta_fotos_pdf,  # ← foto do PDF
            )
            if arq_resp:
                resultados["resposta"].append(arq_resp)

        except Exception as e:
            logger.error(f"  ERRO em {arq.name}: {e}", exc_info=True)
            resultados["erros"].append(arq.name)

    total_k = len(resultados["kria"])
    total_r = len(resultados["resposta"])
    total_e = len(resultados["erros"])
    logger.info(
        f"\nMódulo 02 concluído: "
        f"{total_k} Kria, {total_r} resposta(s), {total_e} erro(s)."
    )
    if callback_progresso:
        callback_progresso(len(arquivos), len(arquivos), "Módulo 02 concluído.")

    return resultados


def executar_kria_resposta_de_lista(
    ncs: list,
    nome_base: str,
    relatorio: str,
    modelo_kria: "Path | None" = None,
    pasta_saida_kria: "Path | None" = None,
    modelo_resposta: "Path | None" = None,
    pasta_saida_resp: "Path | None" = None,
    pasta_fotos_nc: "Path | None" = None,
    pasta_fotos_pdf: "Path | None" = None,
) -> dict:
    """
    Gera Kria (Saída A) e Resposta (Saída B) a partir de uma lista de NCs
    (dict com codigo, data_con, data_reparo, tipo_nc, rod_codigo, rod_tag,
    sentido, km_i, km_f, num_foto, prazo_dias). Usado pelo pipeline Meio Ambiente
    (equivalente M2 a partir do PDF, sem planilha EAF).
    Retorna {"kria": Path | None, "resposta": Path | None}.
    """
    modelo_kria = modelo_kria or M02_MODELO_KRIA
    pasta_saida_kria = pasta_saida_kria or M02_SALVAR_FOTO
    modelo_resposta = modelo_resposta or M02_MODELO_RESP
    pasta_saida_resp = pasta_saida_resp or M02_PENDENTES
    pasta_fotos_nc = pasta_fotos_nc or M02_FOTOS_NC
    pasta_fotos_pdf = pasta_fotos_pdf or M02_FOTOS_PDF

    if not ncs:
        logger.warning("executar_kria_resposta_de_lista: lista de NCs vazia.")
        return {"kria": None, "resposta": None}

    arq_kria = _gerar_kria(
        ncs, nome_base,
        modelo_kria, pasta_saida_kria,
        pasta_fotos_nc,
        relatorio,
        pasta_fotos_pdf_fallback=pasta_fotos_pdf,
    )
    arq_resp = _gerar_resposta(
        ncs, modelo_resposta, pasta_saida_resp,
        pasta_fotos_pdf,
    )
    return {"kria": arq_kria, "resposta": arq_resp}
